"""
K-line incremental fetch + Notion sync + Excel export.

Data source: yfinance (Yahoo Finance) — works from any IP including GitHub Actions.
Taiwan stocks: {SID}.TW for TWSE-listed, {SID}.TWO for TPEX OTC.

Logic:
  1. Read batch_seq*.csv from last 30 days → all scanned stock IDs
  2. Load kline_cache.json → find global latest cached date (incremental)
  3. Batch-download via yfinance (.TW first, .TWO for misses)
  4. Merge new bars into cache (append, dedup by date, sort)
  5. Save updated output/kline_cache.json
  6. Sync 30-day/90-day stats to Notion (PATCH schema first, then update pages)
  7. Export output/kline_export.xlsx (K線匯總 + OHLCV近30日)
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import logging

import pandas as pd
import requests
import yfinance as yf
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Silence yfinance's own noisy warnings ("possibly delisted; no timezone found")
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
logging.getLogger("peewee").setLevel(logging.CRITICAL)

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
SCAN_DIR   = ROOT / "output" / "full_scan"
CACHE_FILE = ROOT / "output" / "kline_cache.json"
EXCEL_FILE = ROOT / "output" / "kline_export.xlsx"

WINDOW_DAYS = 30   # collect stocks scanned within this many days
BATCH_SIZE  = 200  # stocks per yfinance download call

# Lookback calendar days per interval (for first-time / missing stocks)
INTERVAL_LOOKBACK: dict[str, int] = {
    "1d":  90,    # ~60 trading bars
    "1wk": 730,   # ~2 years / ~104 bars
    "1mo": 1825,  # ~5 years / ~60 bars
}

NOTION_API = "https://api.notion.com/v1"
NOTION_VER = "2022-06-28"

# ── Logging setup ──────────────────────────────────────────────────────────────
logger.remove()
logger.add(
    sys.stdout,
    format="<green>{time:HH:mm:ss}</green> | <level>{level:<7}</level> | {message}",
    level="DEBUG",
    colorize=True,
)


# ── Date helpers ───────────────────────────────────────────────────────────────
def today_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def days_ago_iso(n: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=n)).strftime("%Y-%m-%d")


def tomorrow_iso() -> str:
    """yfinance end date is exclusive — pass tomorrow to include today."""
    return (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")


# ── Step 1: Collect stock IDs from last 30 days of scan CSVs ──────────────────
def get_stock_ids_last_30_days() -> list[str]:
    if not SCAN_DIR.exists():
        logger.warning(f"Scan directory not found: {SCAN_DIR}")
        return []

    cutoff = days_ago_iso(WINDOW_DAYS)
    files: list[Path] = []
    for f in sorted(SCAN_DIR.iterdir()):
        if not f.is_file():
            continue
        m = re.match(r"^batch_seq\d+_(\d{4}-\d{2}-\d{2})\.csv$", f.name)
        if m and m.group(1) >= cutoff:
            files.append(f)

    if not files:
        logger.warning("No batch_seq CSV files in last 30 days — trying all available")
        all_files = sorted(
            [f for f in SCAN_DIR.iterdir()
             if re.match(r"^batch_seq\d+_\d{4}-\d{2}-\d{2}\.csv$", f.name)],
            reverse=True,
        )
        files = all_files[:9]

    ids: set[str] = set()
    dates_seen: set[str] = set()
    for f in files:
        try:
            df = pd.read_csv(f, usecols=["stock_id"], encoding="utf-8-sig", dtype=str)
            ids.update(df["stock_id"].dropna().str.strip().tolist())
            mm = re.search(r"(\d{4}-\d{2}-\d{2})", f.name)
            if mm:
                dates_seen.add(mm.group(1))
        except Exception as exc:
            logger.warning(f"Could not read {f.name}: {exc}")

    logger.info(f"Scan dates found: {', '.join(sorted(dates_seen))}")
    logger.info(f"Unique stocks from last {WINDOW_DAYS} days: {len(ids)}")
    return sorted(ids)


# ── Step 3: yfinance batch download ───────────────────────────────────────────
def _download_batch(
    stock_ids: list[str],
    suffix: str,
    start: str,
    end: str,
    interval: str = "1d",
) -> dict[str, list[dict]]:
    """Download one batch of stocks with a given suffix (.TW or .TWO)."""
    if not stock_ids:
        return {}

    yf_tickers = [f"{sid}{suffix}" for sid in stock_ids]

    try:
        if len(yf_tickers) == 1:
            raw = yf.download(
                yf_tickers[0],
                start=start,
                end=end,
                interval=interval,
                auto_adjust=True,
                progress=False,
            )
            if raw.empty:
                return {}
            # Wrap single-ticker result into the same shape as multi-ticker
            ticker_dfs: dict[str, pd.DataFrame] = {yf_tickers[0]: raw}
        else:
            raw = yf.download(
                yf_tickers,
                start=start,
                end=end,
                interval=interval,
                auto_adjust=True,
                group_by="ticker",
                progress=False,
                threads=True,
            )
            if raw.empty:
                return {}
            ticker_dfs = {}
            # raw.columns is a MultiIndex (ticker, field)
            top_level = raw.columns.get_level_values(0).unique().tolist()
            for t in top_level:
                ticker_dfs[t] = raw[t]
    except Exception as exc:
        logger.warning(f"yfinance batch error ({suffix}): {exc}")
        return {}

    result: dict[str, list[dict]] = {}
    for yf_ticker, df in ticker_dfs.items():
        sid = yf_ticker[: -len(suffix)]   # strip suffix to get clean stock_id
        if df is None or df.empty:
            continue
        df = df.dropna(subset=["Close"])
        if df.empty:
            continue
        bars: list[dict] = []
        for dt_idx, row in df.iterrows():
            dt_str = dt_idx.strftime("%Y-%m-%d") if hasattr(dt_idx, "strftime") else str(dt_idx)[:10]
            try:
                bars.append({
                    "time":   dt_str,
                    "open":   round(float(row["Open"]),  2),
                    "high":   round(float(row["High"]),  2),
                    "low":    round(float(row["Low"]),   2),
                    "close":  round(float(row["Close"]), 2),
                    "volume": int(row["Volume"]) if not pd.isna(row["Volume"]) else 0,
                })
            except Exception:
                continue
        if bars:
            result[sid] = bars

    return result


def fetch_all_klines_yfinance(
    stock_ids: list[str],
    start: str,
    end: str,
    interval: str = "1d",
) -> dict[str, list[dict]]:
    """
    Download K-line data for Taiwan stocks via yfinance.
    Phase 1: try {sid}.TW  (TWSE-listed stocks)
    Phase 2: try {sid}.TWO (TPEX OTC) for any stocks not found in phase 1
    Returns {stock_id: [bar_dict, ...]}
    """
    result: dict[str, list[dict]] = {}

    # ── Phase 1: .TW (TWSE main board) ───────────────────────────────────────
    tw_misses: list[str] = []
    batches = [stock_ids[i:i + BATCH_SIZE] for i in range(0, len(stock_ids), BATCH_SIZE)]
    logger.info(
        f"Phase 1 (.TW) [{interval}]: {len(stock_ids)} stocks in {len(batches)} batches "
        f"[{start} → {end}]"
    )
    for bi, chunk in enumerate(batches, 1):
        found = _download_batch(chunk, ".TW", start, end, interval=interval)
        result.update(found)
        miss = [sid for sid in chunk if sid not in found]
        tw_misses.extend(miss)
        logger.info(
            f"  .TW batch {bi}/{len(batches)}: {len(found)}/{len(chunk)} found "
            f"({len(miss)} misses)"
        )

    # ── Phase 2: .TWO (TPEX OTC) for misses ──────────────────────────────────
    if tw_misses:
        two_batches = [tw_misses[i:i + BATCH_SIZE] for i in range(0, len(tw_misses), BATCH_SIZE)]
        logger.info(
            f"Phase 2 (.TWO) [{interval}]: {len(tw_misses)} missed stocks in {len(two_batches)} batches"
        )
        for bi, chunk in enumerate(two_batches, 1):
            found = _download_batch(chunk, ".TWO", start, end, interval=interval)
            result.update(found)
            still_miss = len(chunk) - len(found)
            logger.info(
                f"  .TWO batch {bi}/{len(two_batches)}: {len(found)}/{len(chunk)} found "
                f"({still_miss} not found anywhere)"
            )

    logger.info(
        f"yfinance [{interval}] complete: {len(result)}/{len(stock_ids)} stocks returned data"
    )
    return result


# ── Step 5: Compute statistics ────────────────────────────────────────────────
def compute_stats(bars: list[dict], cutoff_30d: str) -> dict[str, Any] | None:
    if not bars:
        return None
    latest = bars[-1]
    bars_30 = [b for b in bars if b.get("time", "") >= cutoff_30d]
    first_30 = bars_30[0] if bars_30 else None
    first_90 = bars[0]

    def pct(new_close: float, old_close: float) -> float | None:
        if not old_close:
            return None
        return round((new_close - old_close) / old_close * 100, 2)

    return {
        "latest_close": latest["close"],
        "latest_date":  latest["time"],
        "return_30d":   pct(latest["close"], first_30["close"]) if first_30 else None,
        "high_30d":     round(max(b["high"] for b in bars_30), 2) if bars_30 else None,
        "low_30d":      round(min(b["low"]  for b in bars_30), 2) if bars_30 else None,
        "return_90d":   pct(latest["close"], first_90["close"]) if len(bars) >= 10 else None,
        "days":         len(bars),
    }


# ── Step 6: Excel export ──────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="微軟正黑體")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")


def _style_header(ws, headers: list[str], col_widths: list[int]) -> None:
    ws.row_dimensions[1].height = 22
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = w


def export_excel(kline_map: dict[str, list], path: Path) -> None:
    cutoff_30d = days_ago_iso(30)
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "K線匯總"
    headers1 = ["股票代號", "最新收盤", "最後更新", "30日漲幅%", "90日漲幅%", "30日最高", "30日最低", "資料天數"]
    widths1  = [10, 10, 14, 12, 12, 10, 10, 10]
    _style_header(ws1, headers1, widths1)

    rows_summary = []
    for sid, bars in kline_map.items():
        s = compute_stats(bars, cutoff_30d)
        if not s:
            continue
        rows_summary.append((sid, s["latest_close"], s["latest_date"],
                             s["return_30d"], s["return_90d"],
                             s["high_30d"], s["low_30d"], s["days"]))
    rows_summary.sort(key=lambda r: (r[3] or -9999), reverse=True)

    for row_idx, row in enumerate(rows_summary, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

    # Sheet 2: OHLCV (last 30 days)
    ws2 = wb.create_sheet("OHLCV(近30日)")
    headers2 = ["股票代號", "日期", "開盤", "最高", "最低", "收盤", "成交量"]
    widths2  = [10, 12, 10, 10, 10, 10, 14]
    _style_header(ws2, headers2, widths2)

    raw_rows: list[tuple] = []
    for sid, bars in kline_map.items():
        for b in bars:
            if b.get("time", "") >= cutoff_30d:
                raw_rows.append((sid, b["time"], b["open"], b["high"],
                                 b["low"], b["close"], b["volume"]))
    raw_rows.sort(key=lambda r: (r[0], r[1]))

    for row_idx, row in enumerate(raw_rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info(f"Excel saved: {path.name}  ({len(rows_summary)} stocks, {len(raw_rows)} OHLCV rows)")


# ── Step 7: Notion sync ────────────────────────────────────────────────────────
def _notion_headers(token: str) -> dict[str, str]:
    return {
        "Authorization":  f"Bearer {token}",
        "Notion-Version": NOTION_VER,
        "Content-Type":   "application/json",
    }


NOTION_SYNC_TIMEOUT = 480   # seconds (8 minutes)


def sync_to_notion(kline_map: dict[str, list], token: str, db_id: str) -> None:
    if not token or not db_id:
        logger.info("Notion: skipped (NOTION_TOKEN or NOTION_DATABASE_ID not set)")
        return

    cutoff_30d = days_ago_iso(30)
    deadline = time.time() + NOTION_SYNC_TIMEOUT

    # Ensure K-line properties exist in the database schema
    try:
        requests.patch(
            f"{NOTION_API}/databases/{db_id}",
            headers=_notion_headers(token),
            json={"properties": {
                "30日漲幅%": {"number": {"format": "percent"}},
                "30日最高":  {"number": {"format": "number"}},
                "30日最低":  {"number": {"format": "number"}},
                "90日漲幅%": {"number": {"format": "percent"}},
                "K線更新日": {"rich_text": {}},
            }},
            timeout=25,
        ).raise_for_status()
        logger.info("Notion schema patch: OK")
    except Exception as exc:
        logger.warning(f"Notion schema patch warning: {exc}")

    # Build a set of stock_ids we still need to update (stop early once all done)
    remaining = set(kline_map.keys())
    cursor: str | None = None
    updated = failed = total = 0
    MAX_PAGES = 300   # safety cap: stop after scanning 30,000 Notion pages

    while remaining:
        if time.time() >= deadline:
            logger.warning(
                f"Notion: {NOTION_SYNC_TIMEOUT}s timeout reached — "
                f"{updated} updated, {len(remaining)} skipped (will catch up next run)"
            )
            break
        if total // 100 >= MAX_PAGES:
            logger.warning(f"Notion: reached {MAX_PAGES}-page cap, stopping early")
            break

        body: dict = {"page_size": 100}
        if cursor:
            body["start_cursor"] = cursor
        try:
            resp = requests.post(
                f"{NOTION_API}/databases/{db_id}/query",
                headers=_notion_headers(token),
                json=body,
                timeout=25,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as exc:
            logger.error(f"Notion query error: {exc}")
            break

        pages = data.get("results", [])
        total += len(pages)

        for page in pages:
            rt = page.get("properties", {}).get("股票代號", {}).get("rich_text", [])
            sid = rt[0]["text"]["content"].strip() if rt else ""
            if not sid or sid not in remaining:
                continue

            stats = compute_stats(kline_map[sid], cutoff_30d)
            if not stats:
                remaining.discard(sid)
                continue

            props: dict = {
                "30日最高":  {"number": stats["high_30d"]},
                "30日最低":  {"number": stats["low_30d"]},
                "K線更新日": {"rich_text": [{"text": {"content": stats["latest_date"]}}]},
            }
            if stats["return_30d"] is not None:
                props["30日漲幅%"] = {"number": stats["return_30d"] / 100}
            if stats["return_90d"] is not None:
                props["90日漲幅%"] = {"number": stats["return_90d"] / 100}

            try:
                requests.patch(
                    f"{NOTION_API}/pages/{page['id']}",
                    headers=_notion_headers(token),
                    json={"properties": props},
                    timeout=25,
                ).raise_for_status()
                updated += 1
                remaining.discard(sid)
            except Exception as exc:
                logger.debug(f"  Notion update {sid} failed: {exc}")
                failed += 1

        if not data.get("has_more"):
            break
        cursor = data["next_cursor"]

        if updated % 100 == 0 and updated > 0:
            logger.info(f"  Notion progress: {updated} updated, {len(remaining)} remaining")

        if data.get("has_more"):
            cursor = data["next_cursor"]
        else:
            break

    logger.info(
        f"Notion: scanned {total} pages, updated {updated}"
        + (f", {failed} failed" if failed else "")
    )


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    logger.info("=== K-line incremental fetch (yfinance) ===")

    # ── 1. Collect stock IDs from last 30 days of scan CSVs ──────────────────
    stock_ids = get_stock_ids_last_30_days()
    if not stock_ids:
        logger.warning("No stocks found in last 30 days of scans — exiting")
        sys.exit(0)

    # ── 2. Load existing cache (support both old list format and new dict format) ─
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    # New format: {stock_id: {"1d": [...], "1wk": [...], "1mo": [...]}}
    # Old format: {stock_id: [...]}  ← auto-migrated on load
    existing_cache: dict[str, dict[str, list]] = {}
    if CACHE_FILE.exists():
        try:
            raw = json.loads(CACHE_FILE.read_text("utf-8"))
            migrated = 0
            for sid, val in raw.items():
                if isinstance(val, list):
                    existing_cache[sid] = {"1d": val, "1wk": [], "1mo": []}
                    migrated += 1
                else:
                    existing_cache[sid] = val
            logger.info(
                f"Cache loaded: {len(existing_cache)} stocks"
                + (f" ({migrated} migrated from old format)" if migrated else "")
            )
        except Exception as exc:
            logger.warning(f"Failed to load cache ({exc}) — starting fresh")

    # ── 3 & 4. Fetch each interval with per-interval incremental logic ────────
    end_date = tomorrow_iso()
    # Accumulate all fetched bars: {stock_id: {interval: [bars]}}
    all_fetched: dict[str, dict[str, list]] = {}

    for ivl, lookback in INTERVAL_LOOKBACK.items():
        logger.info(f"=== Fetching {ivl} bars ===")

        # Get per-stock cached bars for this interval
        cached_ivl = {
            sid: existing_cache.get(sid, {}).get(ivl, [])
            for sid in stock_ids
        }
        # Stocks with too few bars are treated as missing (force full backfill)
        MIN_BARS = {"1d": 30, "1wk": 10, "1mo": 3}
        threshold = MIN_BARS.get(ivl, 2)
        missing_ids = [sid for sid in stock_ids if len(cached_ivl[sid]) < threshold]
        present_ids = [sid for sid in stock_ids if len(cached_ivl[sid]) >= threshold]

        new_ivl: dict[str, list] = {}

        if missing_ids:
            full_start = days_ago_iso(lookback)
            logger.info(f"  Full backfill ({ivl}): {len(missing_ids)} stocks from {full_start}")
            new_ivl.update(fetch_all_klines_yfinance(missing_ids, full_start, end_date, interval=ivl))

        if present_ids:
            latest = max(
                (cached_ivl[sid][-1]["time"] for sid in present_ids if cached_ivl[sid]),
                default="",
            )
            if latest:
                inc_start = (date.fromisoformat(latest) + timedelta(days=1)).isoformat()
                if inc_start < end_date:
                    logger.info(f"  Incremental ({ivl}): {len(present_ids)} stocks from {inc_start}")
                    new_ivl.update(fetch_all_klines_yfinance(present_ids, inc_start, end_date, interval=ivl))
                else:
                    logger.info(f"  {ivl} already up-to-date (cached through {latest})")

        for sid, bars in new_ivl.items():
            all_fetched.setdefault(sid, {})[ivl] = bars

    # ── 5. Merge fetched bars into kline_map ──────────────────────────────────
    kline_map: dict[str, dict[str, list]] = {}
    all_sids = set(list(existing_cache.keys()) + stock_ids)
    total_new_bars = 0

    for sid in all_sids:
        kline_map[sid] = {}
        for ivl in INTERVAL_LOOKBACK:
            old_bars = existing_cache.get(sid, {}).get(ivl, [])
            new_bars = all_fetched.get(sid, {}).get(ivl, [])
            existing_times = {b["time"] for b in old_bars}
            to_add = [b for b in new_bars if b["time"] not in existing_times]
            merged = old_bars + to_add
            merged.sort(key=lambda b: b.get("time", ""))
            kline_map[sid][ivl] = merged
            total_new_bars += len(to_add)

    covered = sum(1 for sid in stock_ids if kline_map.get(sid, {}).get("1d"))
    logger.info(
        f"Merge complete: +{total_new_bars} new bars, "
        f"{covered}/{len(stock_ids)} scanned stocks have daily data"
    )

    # ── 6. Save updated cache ─────────────────────────────────────────────────
    CACHE_FILE.write_text(json.dumps(kline_map, ensure_ascii=False), "utf-8")
    logger.info(f"Cache saved: {CACHE_FILE.name} ({len(kline_map)} stocks, 3 intervals)")

    # Flatten to daily bars for Excel + Notion (they don't need weekly/monthly)
    kline_map_1d = {sid: data.get("1d", []) for sid, data in kline_map.items()}

    # ── 7. Excel export ───────────────────────────────────────────────────────
    logger.info("=== Excel export ===")
    try:
        export_excel(kline_map_1d, EXCEL_FILE)
    except Exception as exc:
        logger.error(f"Excel export failed: {exc}")

    # ── 8. Notion sync ────────────────────────────────────────────────────────
    logger.info("=== Notion sync ===")
    notion_token = os.environ.get("NOTION_TOKEN",       "").strip()
    notion_db_id = os.environ.get("NOTION_DATABASE_ID", "").strip()
    try:
        sync_to_notion(kline_map_1d, notion_token, notion_db_id)
    except Exception as exc:
        logger.error(f"Notion sync failed: {exc}")

    logger.info("=== Done ===")


if __name__ == "__main__":
    main()
