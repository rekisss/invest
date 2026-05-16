"""
backtest_sweep.py – Parameter grid search for the Taiwan MACD strategy.

Loads price/institutional data ONCE (FinMindClient caches to disk), then
evaluates every combination in PARAM_GRID and writes a ranked summary.

Usage:
    python backtest_sweep.py \
        --start 2020-01-01 --end 2024-12-31 \
        --stocks auto --max-universe 80 \
        --output output/sweep

Edit PARAM_GRID below to define the search space.
"""
from __future__ import annotations

import argparse
import itertools
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import pandas as pd

_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())

from data_loader import (
    FinMindClient,
    fetch_institutional_data,
    fetch_market_index,
    fetch_stock_info,
    fetch_stock_prices,
    load_stock_list,
)
from strategy import StrategyConfig, prepare_market_frame, prepare_stock_signals
from backtest import run_backtest
from universe import build_auto_universe

# ── Edit this grid to define the sweep space ──────────────────────────────────
PARAM_GRID: dict[str, list[Any]] = {
    "rsi_threshold":   [50.0, 55.0, 60.0],
    "stop_loss_pct":   [0.04, 0.05, 0.06],
    "take_profit_pct": [0.08, 0.10, 0.12],
}

# These stay fixed across all sweep runs
FIXED_PARAMS: dict[str, Any] = {
    "trailing_stop_pct": 0.07,
    "max_positions": 3,
    "adx_threshold": 20.0,
    "volume_multiplier": 1.5,
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Backtest parameter grid sweep.")
    p.add_argument("--start", default="2020-01-01")
    p.add_argument("--end", default=None, help="End date (default: today CST).")
    p.add_argument("--stocks", default="auto")
    p.add_argument("--max-universe", type=int, default=80)
    p.add_argument("--capital", type=float, default=1_000_000)
    p.add_argument("--output", default="output/sweep")
    p.add_argument("--workers", type=int, default=8)
    return p.parse_args()


def _fetch_all_raw(
    stock_list: pd.DataFrame,
    client: FinMindClient,
    start: str,
    end: str,
    workers: int,
) -> dict[str, dict]:
    """Fetch price + institutional data for every stock into memory."""
    records = stock_list.to_dict("records")
    total = len(records)
    raw: dict[str, dict] = {}

    def _load(stock: dict) -> tuple[str, dict] | None:
        try:
            prices = fetch_stock_prices(client, stock["stock_id"], start, end)
            if prices.empty:
                return None
            inst = fetch_institutional_data(client, stock["stock_id"], start, end)
            return stock["stock_id"], {"info": stock, "prices": prices, "inst": inst}
        except Exception as exc:
            print(f"[warn] {stock['stock_id']}: {exc}")
            return None

    done = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_load, s) for s in records]
        for fut in as_completed(futures):
            done += 1
            if done % 10 == 0 or done == total:
                print(f"[data] {done}/{total} loaded", flush=True)
            res = fut.result()
            if res:
                raw[res[0]] = res[1]
    return raw


def _run_one(
    combo: dict[str, Any],
    raw: dict[str, dict],
    market_precomputed: pd.DataFrame,
    capital: float,
) -> dict[str, Any]:
    """Compute signals + backtest for a single parameter combination.

    market_precomputed should already have had prepare_market_frame applied with
    default market_ma_window=60.  Since we never sweep market_ma_window, we avoid
    recomputing it for every combination.
    """
    config = StrategyConfig(**{**FIXED_PARAMS, **combo})
    mkt = market_precomputed
    signals: dict[str, pd.DataFrame] = {}
    for stock_id, data in raw.items():
        try:
            frame = prepare_stock_signals(
                data["info"], data["prices"], mkt, data["inst"], config
            )
            signals[stock_id] = frame
        except Exception:
            pass
    if not signals:
        return {**combo, "total_return_pct": float("nan"), "sharpe_ratio": float("nan"),
                "max_drawdown_pct": float("nan"), "win_rate_pct": float("nan"),
                "total_trades": 0, "alpha_pct": float("nan")}
    m = run_backtest(signals, mkt, config, capital)["metrics"]
    return {
        **combo,
        "total_return_pct": m.get("total_return_pct", float("nan")),
        "sharpe_ratio":     m.get("sharpe_ratio",     float("nan")),
        "max_drawdown_pct": m.get("max_drawdown_pct", float("nan")),
        "win_rate_pct":     m.get("win_rate_pct",     float("nan")),
        "total_trades":     m.get("total_trades",     0),
        "alpha_pct":        m.get("alpha_pct",        float("nan")),
    }


def main() -> None:
    args = parse_args()
    end = args.end or pd.Timestamp.now().strftime("%Y-%m-%d")
    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    client = FinMindClient(cache_dir=out_dir / "cache")

    if args.stocks != "auto":
        stock_list = load_stock_list(args.stocks)
    else:
        stock_list = build_auto_universe(fetch_stock_info(client), max_symbols=args.max_universe)

    print(f"[sweep] {len(stock_list)} stocks · {args.start} → {end}")

    market_raw = fetch_market_index(client, args.start, end)
    if market_raw.empty:
        raise RuntimeError("Cannot fetch TAIEX — check FINMIND_TOKEN.")

    print("[sweep] Fetching stock data (cached after first run)…")
    raw = _fetch_all_raw(stock_list, client, args.start, end, args.workers)
    print(f"[sweep] {len(raw)} stocks ready")

    # Pre-compute market frame once — market_ma_window is not swept
    base_config = StrategyConfig(**FIXED_PARAMS)
    market_precomputed = prepare_market_frame(market_raw, base_config)

    keys = list(PARAM_GRID.keys())
    combos = [dict(zip(keys, vals)) for vals in itertools.product(*PARAM_GRID.values())]
    print(f"[sweep] Running {len(combos)} combinations…\n")

    rows: list[dict] = []
    for i, combo in enumerate(combos, 1):
        label = "  ".join(f"{k}={v}" for k, v in combo.items())
        print(f"[{i:>3}/{len(combos)}] {label}", flush=True)
        rows.append(_run_one(combo, raw, market_precomputed, args.capital))

    df = pd.DataFrame(rows).sort_values("sharpe_ratio", ascending=False).reset_index(drop=True)

    csv_path = out_dir / "sweep_results.csv"
    df.to_csv(csv_path, index=False)

    xlsx_path = out_dir / "sweep_results.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="sweep")
        ws = writer.sheets["sweep"]
        from openpyxl.styles import PatternFill, Font
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        bold = Font(bold=True)
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=1, column=col_idx).font = bold
        for row_idx in range(2, min(7, len(df) + 2)):
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = green
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    print(f"\n[sweep] Done → {xlsx_path}")
    print("\nTop 5 by Sharpe ratio:")
    print(df.head(5).to_string(index=False))

    best = df.iloc[0]
    print("\n★ Best combination:")
    for k in keys:
        print(f"   {k}: {best[k]}")
    print(f"   Sharpe: {best['sharpe_ratio']:.3f}  Return: {best['total_return_pct']:.1f}%  "
          f"MaxDD: {best['max_drawdown_pct']:.1f}%  WinRate: {best['win_rate_pct']:.1f}%")


if __name__ == "__main__":
    main()
