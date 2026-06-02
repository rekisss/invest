from __future__ import annotations

import argparse
from collections import deque
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import os
from pathlib import Path
import random
import sys
import time
from typing import Any

# 自動載入專案目錄的 .env 檔案（若存在）
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())

from datetime import datetime, timezone, timedelta

import pandas as pd
import numpy as np

from backtest import run_backtest
from data_loader import (
    FinMindClient,
    clean_cache,
    fetch_all_margin_data,
    fetch_all_monthly_revenue,
    fetch_disposition_stocks,
    fetch_financial_statement_dates,
    fetch_fundamentals,
    fetch_institutional_data,
    fetch_market_index,
    fetch_market_institutional,
    fetch_market_margin,
    fetch_options_pcr,
    fetch_stock_info,
    fetch_stock_kbar,
    fetch_stock_prices,
    fetch_suspended_stocks,
    load_stock_list,
    probe_batch_quota,
    validate_finmind_token,
)
from fundamentals import compute_f_score
from fugle_client import FugleClient, fetch_watch_quotes
from news_service import NewsClient, summarize_news, fetch_market_news_sentiment, format_market_news_block
from notifier import send_discord_messages, split_message
from report import save_hybrid_report, save_reports, save_scan_report, save_sponsor_monitor_report
from strategy import (
    StrategyConfig,
    compute_market_breadth,
    compute_market_regime,
    latest_signal_snapshot,
    prepare_market_frame,
    prepare_stock_signals,
    rank_candidates,
)
from universe import build_auto_universe
from notion_sync import confidence_score as _confidence_score, notion_enabled, recommend_observation_period, sync_scan_results
from market_predictor import (
    MarketPredictor, fetch_us_features, format_prediction_block,
    generate_scenario_analysis, format_us_block,
)
from taiwan_futures import (
    fetch_futures_daily, fetch_futures_institutional,
    build_futures_features, format_futures_block,
    fetch_night_session, format_night_session_block,
)
from economic_calendar import fetch_upcoming_events, format_calendar_block

_CST = timezone(timedelta(hours=8))


def _cst_now(fmt: str = "%H:%M") -> str:
    return datetime.now(_CST).strftime(fmt)


def _cst_today() -> str:
    """Return today's date in CST (UTC+8) as YYYY-MM-DD."""
    return _cst_now("%Y-%m-%d")


def _write_batch_markdown(
    batch_dir: "Path",
    batch_index: int,
    batch_count: int,
    scan_date: str,
    candidates: "pd.DataFrame",
    watchlist: "pd.DataFrame",
    breadth: dict,
) -> None:
    """Write a human-readable Markdown summary of a completed batch scan."""
    md_path = batch_dir / f"batch_{batch_index:02d}.md"
    lines: list[str] = []
    lines.append(f"# 批次 {batch_index}/{batch_count - 1} · {scan_date} {_cst_now()} CST\n")
    lines.append(
        f"掃描：**{breadth.get('total_stocks', 0)}** 檔｜"
        f"候選：**{len(candidates)}** 檔｜"
        f"觀察：**{len(watchlist)}** 檔\n"
    )

    def _tbl_row(r: dict, cols: list[str]) -> str:
        return "| " + " | ".join(str(r.get(c, "")) for c in cols) + " |"

    if not candidates.empty:
        lines.append("## 候選")
        show_cols = ["stock_id", "name", "close", "entry_score", "condition_count", "volume_ratio", "entry_reason"]
        show_cols = [c for c in show_cols if c in candidates.columns]
        header_map = {
            "stock_id": "股票", "name": "名稱", "close": "收盤",
            "entry_score": "評分", "condition_count": "條件數",
            "volume_ratio": "量比", "entry_reason": "訊號",
        }
        headers = [header_map.get(c, c) for c in show_cols]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join("---" for _ in headers) + "|")
        for _, row in candidates.head(30).iterrows():
            vals = []
            for c in show_cols:
                v = row.get(c, "")
                if c in ("entry_score", "volume_ratio") and v != "":
                    try:
                        v = f"{float(v):.2f}"
                    except Exception:
                        pass
                elif c == "close" and v != "":
                    try:
                        v = f"{float(v):.1f}"
                    except Exception:
                        pass
                vals.append(str(v))
            lines.append("| " + " | ".join(vals) + " |")
        lines.append("")

    if not watchlist.empty:
        lines.append("## 觀察清單")
        watch_cols = ["stock_id", "name", "close", "entry_score", "condition_count", "skip_reason"]
        watch_cols = [c for c in watch_cols if c in watchlist.columns]
        w_header_map = {
            "stock_id": "股票", "name": "名稱", "close": "收盤",
            "entry_score": "評分", "condition_count": "條件數", "skip_reason": "未進候選原因",
        }
        w_headers = [w_header_map.get(c, c) for c in watch_cols]
        lines.append("| " + " | ".join(w_headers) + " |")
        lines.append("|" + "|".join("---" for _ in w_headers) + "|")
        for _, row in watchlist.head(20).iterrows():
            vals = []
            for c in watch_cols:
                v = row.get(c, "")
                if c in ("entry_score",) and v != "":
                    try:
                        v = f"{float(v):.2f}"
                    except Exception:
                        pass
                elif c == "close" and v != "":
                    try:
                        v = f"{float(v):.1f}"
                    except Exception:
                        pass
                vals.append(str(v))
            lines.append("| " + " | ".join(vals) + " |")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    _safe_print(f"[batch {batch_index}] Markdown 摘要 → {md_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Taiwan MACD swing strategy backtester and scanner.")
    parser.add_argument("--mode", choices=["backtest", "scan", "hybrid-monitor", "sponsor-monitor", "event-monitor", "daily-report", "walk-forward", "predict", "aggregate", "partial-aggregate", "continue-scan", "smart-scan", "fill-batch-gaps", "fill-gaps", "sequential-scan"], default="scan")
    parser.add_argument("--stocks", default="auto", help="CSV file containing stock_id and optional name, or 'auto'.")
    parser.add_argument("--start", default="2020-01-01")
    parser.add_argument("--end", default=_cst_today())
    parser.add_argument("--capital", type=float, default=1_000_000)
    parser.add_argument("--output", default="output")
    parser.add_argument("--lookback-days", type=int, default=420)
    parser.add_argument("--max-universe", type=int, default=120)
    parser.add_argument("--top-n", type=int, default=20)
    parser.add_argument("--watch-top", type=int, default=10, help="Maximum live-watch symbols.")
    parser.add_argument("--interval-seconds", type=int, default=120, help="Polling interval for repeated monitor runs.")
    parser.add_argument("--repeat-count", type=int, default=1, help="How many monitor cycles to run.")
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--max-price", type=float, default=None, help="Optional maximum close price filter for ranking.")
    parser.add_argument("--prefer-lower-price", action="store_true", help="Prefer lower-priced names when ranking candidates.")
    parser.add_argument("--include-news", action="store_true", help="Attach recent news summary to notifications.")
    parser.add_argument("--news-limit", type=int, default=2, help="Maximum number of news headlines per stock in notifications.")
    parser.add_argument("--event-rise-threshold", type=float, default=0.045, help="Intraday rise threshold for event monitor.")
    parser.add_argument("--event-drop-threshold", type=float, default=-0.035, help="Intraday drop threshold for event monitor.")
    parser.add_argument("--event-volume-multiplier", type=float, default=2.5, help="Volume surge multiple versus previous sample.")
    parser.add_argument("--event-cooldown-seconds", type=int, default=600, help="Cooldown before the same symbol+event can notify again.")
    parser.add_argument("--notify", action="store_true")
    parser.add_argument("--use-earnings-filter", action="store_true")
    parser.add_argument("--next-day-fill", default=True, action=argparse.BooleanOptionalAction, help="Fill backtest entries at next day open (default: True). Use --no-next-day-fill to revert to signal-day close.")
    parser.add_argument("--heartbeat-minutes", type=int, default=0, help="Send a status snapshot every N minutes during event-monitor (0 = disabled).")
    parser.add_argument("--clean-cache", action="store_true", help="Delete stale cache files before running.")
    parser.add_argument("--clean-cache-days", type=int, default=30, help="Age threshold in days for --clean-cache (default 30).")
    parser.add_argument("--wf-folds", type=int, default=4, help="Number of folds for walk-forward analysis (default 4).")
    parser.add_argument("--wf-overlap-days", type=int, default=0, help="Overlap days between walk-forward folds (default 0).")
    parser.add_argument("--watch-extra", default="", help="Comma-separated stock IDs to always include in event-monitor watch list.")
    parser.add_argument("--min-confidence", type=int, default=0, help="Minimum confidence score (0-100) to include a candidate in notifications (default 0 = all).")
    parser.add_argument("--use-atr-stop", action="store_true", help="Use ATR-based stop loss instead of fixed percentage in backtest.")
    parser.add_argument("--atr-stop-multiplier", type=float, default=2.0, help="ATR multiplier for stop loss (default 2.0, used with --use-atr-stop).")
    parser.add_argument("--max-holding-days", type=int, default=0, help="Force-exit positions after N calendar days in backtest (0 = disabled).")
    parser.add_argument("--max-positions-per-sector", type=int, default=2, help="Max simultaneous positions per industry sector in backtest (default 2, 0 = unlimited).")
    parser.add_argument("--f-score-min", type=int, default=0, help="Minimum Piotroski F-Score to qualify for entry (0 = disabled, 6 = recommended).")
    parser.add_argument("--batch-index", type=int, default=-1, help="0-based batch index for full-market scan (-1 = no batching).")
    parser.add_argument("--batch-count", type=int, default=8, help="Total number of batches for full-market scan (default 8).")
    parser.add_argument("--segment-start", type=int, default=-1, help="Explicit start position in sorted universe (for weighted multi-quota splits).")
    parser.add_argument("--segment-size", type=int, default=-1, help="Explicit segment size (number of stocks). Overrides equal --batch-count division.")
    parser.add_argument("--positions", default="positions.csv", help="Path to open positions CSV for monitoring (default positions.csv).")
    # StrategyConfig overrides (for parallel backtest experiments)
    parser.add_argument("--rsi-threshold", type=float, default=None, help="RSI threshold for rsi_strong signal (default 55.0).")
    parser.add_argument("--adx-threshold", type=float, default=None, help="ADX threshold for adx_trending signal (default 20.0).")
    parser.add_argument("--stop-loss-pct", type=float, default=None, help="Fixed stop-loss percentage (default 0.05).")
    parser.add_argument("--take-profit-pct", type=float, default=None, help="Take-profit percentage (default 0.10).")
    parser.add_argument("--trailing-stop-pct", type=float, default=None, help="Trailing stop percentage (default 0.07).")
    parser.add_argument("--volume-multiplier", type=float, default=None, help="Volume expansion multiplier for volume_break signal (default 1.5).")
    parser.add_argument("--max-positions", type=int, default=None, help="Maximum concurrent positions (default 3).")
    return parser.parse_args()


_SENTIMENT_EMOJI: dict[str, str] = {
    "positive": "📈",
    "negative": "📉",
    "neutral": "📰",
}

ENTRY_REASON_LABELS = {
    "macd_golden_cross": "MACD黃金交叉",
    "hist_turn_positive": "柱狀圖翻正",
    "above_ema60": "站上EMA60",
    "ema60_gt_ema120": "EMA60大於EMA120",
    "volume_break": "量能放大",
    "rsi_strong": "RSI偏強",
    "adx_trending": "ADX趨勢成立",
    "breakout_20d": "突破20日高點",
    "market_above_ma60": "大盤站上MA60",
    "foreign_buy_3d": "外資連3買",
    "avoid_chase": "未過度追價",
    "liquidity_ok": "流動性合格",
    "stronger_than_market": "強於大盤",
    "kd_golden_cross": "KD黃金交叉",
    "obv_uptrend": "OBV趨勢向上",
    "invest_trust_buy_2d": "投信連買2天",
    "bb_squeeze_breakout": "BB壓縮突破",
    "breakout_volume_confirm": "突破量能確認",
    "williams_r_recovery": "W%R超賣回升",
    "cci_momentum": "CCI動能強勁",
    "mfi_strong": "MFI資金流入",
    "above_ichimoku_cloud": "站上一目雲",
    "dealer_buy_3d": "自營連買3天",
}

_MAX_CONDITION_COUNT = 23

_HARD_CONDITIONS = frozenset([
    "macd_golden_cross", "hist_turn_positive", "above_ema60", "ema60_gt_ema120",
    "volume_break", "rsi_strong", "breakout_20d", "market_above_ma60",
    "avoid_chase", "liquidity_ok",
])


def _missing_hard_count(entry_reason: object) -> int:
    met = {p.strip() for p in str(entry_reason or "").split(",") if p.strip()}
    return len(_HARD_CONDITIONS - met)


def _missing_hard_labels(entry_reason: object, max_items: int = 3) -> str:
    met = {p.strip() for p in str(entry_reason or "").split(",") if p.strip()}
    missing = _HARD_CONDITIONS - met
    labels = [ENTRY_REASON_LABELS.get(c, c) for c in sorted(missing)]
    return " / ".join(labels[:max_items]) if labels else ""



def _int_or(v: object, default: int = 0) -> int:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _float_or(v: object, default: float = 0.0) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _entry_stop_target(close: float, atr: float | None, stop_pct: float = 0.05, target_pct: float = 0.10) -> tuple[str, str, str, str]:
    if atr and atr > 0:
        # ATR-based: entry within 0.5 ATR, stop 2 ATR below, target 3 ATR above
        entry_hi = round(close + 0.5 * atr, 2)
        stop = round(close - 2.0 * atr, 2)
        target = round(close + 3.0 * atr, 2)
    else:
        entry_hi = round(close * 1.015, 2)
        stop = round(close * (1 - stop_pct), 2)
        target = round(close * (1 + target_pct), 2)
    rr = round((target - close) / max(close - stop, 0.01), 1)
    return f"{close:.2f}–{entry_hi:.2f}", f"{stop:.2f}", f"{target:.2f}", f"{rr}:1"


_EXCEL_SUMMARY_COLS = [
    ("stock_id",            "股票代號"),
    ("name",                "名稱"),
    ("industry_category",   "產業"),
    ("entry_score",         "進場分數"),
    ("close",               "收盤價"),
    ("date",                "資料日期"),
    ("entry_signal",        "進場訊號"),
    ("rsi14",               "RSI14"),
    ("adx14",               "ADX14"),
    ("volume_ratio",        "量比"),
    ("foreign_buy_streak",  "外資連買日"),
    ("invest_trust_streak", "投信連買日"),
    ("f_score",             "F-Score"),
    ("macd_hist",           "MACD直方"),
    ("entry_reason",        "進場原因"),
]


def _write_summary_excel(df: pd.DataFrame, path: "Path") -> None:
    """Write a clean summary Excel: key columns only, sorted by entry_score, green rows for signals."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        src_cols = [c for c, _ in _EXCEL_SUMMARY_COLS if c in df.columns]
        zh_headers = {c: zh for c, zh in _EXCEL_SUMMARY_COLS if c in df.columns}
        out = df[src_cols].copy()
        if "entry_score" in out.columns:
            out = out.sort_values("entry_score", ascending=False)
        out = out.rename(columns=zh_headers)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "掃描結果"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        signal_fill = PatternFill("solid", fgColor="E2EFDA")

        for ci, col_name in enumerate(out.columns, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        has_signal_col = "進場訊號" in out.columns
        for ri, row in enumerate(out.itertuples(index=False), 2):
            has_signal = bool(getattr(row, "進場訊號", False)) if has_signal_col else False
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if has_signal:
                    cell.fill = signal_fill
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as exc:
        _safe_print(f"[excel] summary 寫入失敗（graceful skip）: {exc}")
        df.to_excel(path, index=False, engine="openpyxl")


def _safe_print(message: str = "") -> None:
    try:
        print(message)
    except UnicodeEncodeError:
        encoded = message.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(sys.stdout.encoding or "utf-8", errors="replace")
        print(encoded)


def _reason_labels(raw_reason: object, max_items: int = 3) -> str:
    parts = [part.strip() for part in str(raw_reason or "").split(",") if part.strip()]
    labels = [ENTRY_REASON_LABELS.get(part, part) for part in parts[:max_items]]
    return " / ".join(labels) if labels else "條件接近"


def _low_price_tag(close_value: object) -> str:
    try:
        close_float = float(close_value)
    except (TypeError, ValueError):
        return ""
    if close_float <= 30:
        return "低價"
    if close_float <= 80:
        return "中低價"
    return ""


def build_news_map(
    output_dir: str | Path,
    rows: list[dict[str, object]],
    news_limit: int = 2,
) -> dict[str, dict[str, object]]:
    client = NewsClient(cache_dir=Path(output_dir) / "news_cache")
    news_map: dict[str, dict[str, object]] = {}

    def _fetch_one(row: dict[str, object]) -> tuple[str, dict[str, object]] | None:
        stock_id = str(row.get("stock_id") or "")
        name = str(row.get("name") or stock_id)
        if not stock_id:
            return None
        try:
            frame = client.fetch_stock_news(stock_id, name, limit=news_limit)
            items = frame.to_dict(orient="records")
            return stock_id, {"items": items, "summary": summarize_news(items)}
        except Exception as exc:
            print(f"[news] {stock_id} 抓取失敗: {exc}", file=sys.stderr)
            return None

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(_fetch_one, row) for row in rows[:10]]
        for future in as_completed(futures):
            result = future.result()
            if result:
                stock_id, data = result
                news_map[stock_id] = data

    return news_map


def _news_brief(stock_id: str, news_map: dict[str, dict[str, object]]) -> str:
    payload = news_map.get(stock_id)
    if not payload:
        return ""
    summary = payload.get("summary", {})
    # Prefer recent_sentiment (3-day window) over overall sentiment when fresh news exists
    has_recent = bool(summary.get("has_recent_news"))
    sentiment = str(summary.get("recent_sentiment" if has_recent else "sentiment") or "neutral")
    emoji = _SENTIMENT_EMOJI.get(sentiment, "📰")
    top = summary.get("top_headlines") or []
    titles = [str(h.get("title") or "").replace("\n", " ").strip() for h in top[:2] if h.get("title")]
    if not titles:
        headline = str(summary.get("headline") or "").replace("\n", " ").strip()
        if not headline:
            return ""
        titles = [headline]
    parts = [(t[:38] + "…") if len(t) > 38 else t for t in titles]
    freshness = "" if has_recent else " (舊)"
    return f"\n  {emoji}{freshness} {' / '.join(parts)}"


def _slope_arrow(slope: float, strong: float = 0.30, mild: float = 0.05) -> str:
    if slope > strong:
        return "📈"
    if slope > mild:
        return "↗"
    if slope >= -mild:
        return "→"
    if slope >= -strong:
        return "↘"
    return "📉"


def _slope_label(slope: float, strong: float = 0.30, mild: float = 0.05) -> str:
    if slope > strong:
        return "強勢上升"
    if slope > mild:
        return "緩步上升"
    if slope >= -mild:
        return "橫盤"
    if slope >= -strong:
        return "緩步下跌"
    return "強勢下跌"


def _watchlist_line(row: Any, news_map: dict[str, dict[str, object]]) -> str:
    """Single bullet line for watchlist / near-candidate entries."""
    price_tag = _low_price_tag(row.get("close"))
    price_note = f" `{price_tag}`" if price_tag else ""
    brief = _news_brief(str(row["stock_id"]), news_map)
    missing = _missing_hard_labels(row.get("entry_reason"))
    missing_txt = f" | 缺: {missing}" if missing else ""
    close_val = float(row.get("close") or 0)
    high20 = row.get("close_20d_high")
    gap_txt = ""
    if high20 and pd.notna(high20) and close_val > 0:
        gap_pct = (float(high20) - close_val) / close_val * 100
        if 0 < gap_pct < 5:
            gap_txt = f" | 距突破 `{gap_pct:.1f}%`"
    trend_txt = _trend_label(row)
    return (
        f"• **{row['stock_id']}** {row['name']} | `{_int_or(row['condition_count'])}/{_MAX_CONDITION_COUNT}` | "
        f"收 `{_float_or(row['close']):.2f}`{price_note}{missing_txt}{gap_txt}{trend_txt}{brief}"
    )


def _trend_block(row: Any) -> str:
    """Full trend analysis line for candidate cards: daily + monthly slope + observation estimate."""
    try:
        s20 = float(row.get("lr_slope_20") or 0)
        if pd.isna(s20):
            return ""
    except (TypeError, ValueError):
        return ""
    try:
        s60 = float(row.get("lr_slope_60") or 0)
        s60_ok = not pd.isna(s60)
    except (TypeError, ValueError):
        s60, s60_ok = 0.0, False
    adx = 0.0
    try:
        adx = float(row.get("adx14") or 0)
        if pd.isna(adx):
            adx = 0.0
    except (TypeError, ValueError):
        pass

    d_arrow = _slope_arrow(s20)
    d_desc = _slope_label(s20)
    m_part = f"月線 `{s60:+.2f}%/d` {_slope_arrow(s60, 0.15, 0.03)} {_slope_label(s60, 0.15, 0.03)}" if s60_ok else "月線資料不足"

    both_up = s20 > 0.05 and s60_ok and s60 > 0.03
    both_dn = s20 < -0.05 and s60_ok and s60 < -0.03
    diverging = s60_ok and ((s20 > 0.05 and s60 < -0.03) or (s20 < -0.05 and s60 > 0.03))

    if both_up and adx >= 25:
        obs = "日月線同向強勢，可觀察 **3–4 週**"
    elif both_up and adx >= 20:
        obs = "日月線同向穩定，可觀察 **2–3 週**"
    elif both_up:
        obs = "日月線同向，可觀察 **2 週**"
    elif s20 > 0.05 and not s60_ok:
        obs = "短線上升，可觀察 **1–2 週**"
    elif s20 > 0.05 and not both_up and not diverging:
        obs = "短線反彈但月線偏弱，謹慎 **1–2 週**"
    elif diverging:
        obs = "日月線分歧，趨勢轉折可能，觀察 **1 週**"
    elif both_dn:
        obs = "日月線同向下跌，暫不建議進場"
    elif abs(s20) <= 0.05:
        obs = "橫盤整理，等待突破方向"
    else:
        obs = "觀察 **1–2 週**"

    return f"📊 日線 `{s20:+.2f}%/d` {d_arrow} {d_desc} ｜ {m_part}\n⏳ {obs}"


def _trend_label(row: Any) -> str:
    """Short inline trend for watchlist bullets."""
    try:
        s20 = float(row.get("lr_slope_20") or 0)
        if pd.isna(s20):
            return ""
    except (TypeError, ValueError):
        return ""
    try:
        s60 = float(row.get("lr_slope_60") or 0)
        s60_txt = f"/月`{s60:+.2f}%`" if not pd.isna(s60) else ""
    except (TypeError, ValueError):
        s60_txt = ""
    return f" | {_slope_arrow(s20)} 日`{s20:+.2f}%`{s60_txt}"


def load_universe(args: argparse.Namespace, client: FinMindClient) -> pd.DataFrame:
    if args.stocks != "auto":
        universe = load_stock_list(args.stocks)
    else:
        stock_info = fetch_stock_info(client)
        universe = build_auto_universe(stock_info, max_symbols=args.max_universe)
    batch_index = getattr(args, "batch_index", -1)
    batch_count = max(1, getattr(args, "batch_count", 8))
    if batch_index >= 0 and batch_count > 1:
        n = len(universe)
        size = (n + batch_count - 1) // batch_count
        start = batch_index * size
        universe = universe.iloc[start: start + size].reset_index(drop=True)
    return universe


def collect_signals(
    stock_list: pd.DataFrame,
    client: FinMindClient,
    market: pd.DataFrame,
    config: StrategyConfig,
    start_date: str,
    end_date: str,
    workers: int,
    checkpoint_path: "str | None" = None,
    margin_df: "pd.DataFrame | None" = None,
    revenue_df: "pd.DataFrame | None" = None,
) -> tuple[dict[str, pd.DataFrame], list[pd.DataFrame], list[str], bool]:
    """Returns (signals_by_stock, signal_frames, load_errors, quota_exhausted).

    quota_exhausted is True when a daily-quota 402 was detected mid-scan.
    In that case remaining unfetched stocks are skipped immediately so the
    caller can switch tokens without waiting for hundreds of timeout errors.
    """
    signals_by_stock: dict[str, pd.DataFrame] = {}
    signal_frames: list[pd.DataFrame] = []
    load_errors: list[str] = []

    _req_delay = max(0.0, 1.5 / max(workers, 1))
    _fund_start = (pd.Timestamp(start_date) - pd.Timedelta(days=730)).strftime("%Y-%m-%d")
    _quota_event = threading.Event()  # set when daily quota is confirmed exhausted

    def _is_quota_error(exc: BaseException) -> bool:
        s = str(exc)
        # Only permanent quota exhaustion aborts the scan — transient rate-limits must not.
        return "配額已耗盡" in s or "upper limit" in s.lower()

    def load_one(stock: dict[str, Any]) -> tuple[str, pd.DataFrame] | tuple[str, None]:
        # Skip immediately if quota was already detected as exhausted by another thread
        if _quota_event.is_set():
            return stock["stock_id"], None
        time.sleep(_req_delay + random.uniform(0, _req_delay))
        if _quota_event.is_set():
            return stock["stock_id"], None
        sid = stock["stock_id"]
        try:
            prices = fetch_stock_prices(client, sid, start_date, end_date)
            if prices.empty:
                return sid, None
            institutional = fetch_institutional_data(client, sid, start_date, end_date)
            earnings = pd.DataFrame(columns=["date"])
            if config.use_earnings_filter:
                earnings = fetch_financial_statement_dates(client, sid, start_date, end_date)
            stock_f_score = -1
            if config.f_score_min > 0:
                try:
                    fund = fetch_fundamentals(client, sid, _fund_start, end_date)
                    result = compute_f_score(fund["income"], fund["balance"], fund["cashflow"])
                    stock_f_score = int(result["f_score"])
                except Exception:
                    pass
            frame = prepare_stock_signals(
                stock, prices, market, institutional, config,
                earnings_dates=earnings, f_score=stock_f_score,
                margin_df=margin_df,
                revenue_df=revenue_df,
            )
            return sid, frame
        except Exception as error:
            if _is_quota_error(error):
                _quota_event.set()
            print(f"[warn] skipped {sid}: {error}", file=sys.stderr)
            return sid, error  # type: ignore[return-value]

    _signal_cols = [
        "date", "stock_id", "name", "industry_category",
        "open", "high", "low", "close",
        "entry_signal", "entry_reason", "skip_trade", "skip_reason",
        "base_exit_signal", "base_exit_reason",
        "condition_count", "entry_score",
        "relative_strength_5d", "volume_ratio", "volume_ma20",
        "rsi14", "adx14", "atr14",
        "foreign_buy_streak", "invest_trust_streak",
        "stoch_k", "stoch_d", "bb_pct_b", "bb_bandwidth",
        "obv_uptrend", "bb_squeeze_breakout", "breakout_volume_confirm",
        "dealer_buy_streak", "dealer_buy_3d",
        "williams_r", "cci20", "mfi14", "mfi_strong", "above_ichimoku_cloud",
        "lr_slope_20", "lr_slope_60",
        "f_score", "limit_down_streak",
        "margin_change_5d", "short_ratio",
        "revenue_yoy", "revenue_mom", "revenue_3m_yoy",
    ]
    _signal_cols_present: list[str] = []  # resolved on first result
    _checkpoint_written = False           # tracks whether checkpoint header was written
    stock_records = stock_list.to_dict("records")
    total = len(stock_records)
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(load_one, stock) for stock in stock_records]
        for future in as_completed(futures):
            done += 1
            if done % 10 == 0 or done == total:
                _q = " [配額耗盡，跳過剩餘]" if _quota_event.is_set() else ""
                print(f"[collect_signals] {done}/{total} stocks loaded{_q}", file=sys.stderr)
            result = future.result()
            stock_id, payload = result
            if payload is None:
                continue  # skipped (quota abort or empty prices)
            if isinstance(payload, Exception):
                if not _is_quota_error(payload):
                    load_errors.append(f"{stock_id}: {payload}")
                continue
            signal_frame = payload
            signals_by_stock[stock_id] = signal_frame
            if not _signal_cols_present:
                _signal_cols_present = [c for c in _signal_cols if c in signal_frame.columns]
            signal_frames.append(signal_frame[_signal_cols_present])
            # Incremental checkpoint: write the latest row for this stock immediately.
            # Runs in the main thread (as_completed loop) so no lock needed.
            # This ensures scanned data is preserved even if the process is
            # interrupted mid-scan by quota exhaustion, IP throttle, or crash.
            if checkpoint_path:
                try:
                    _row = signal_frame.iloc[[-1]].copy()
                    _mode = "a" if _checkpoint_written else "w"
                    _row.to_csv(checkpoint_path, mode=_mode, header=not _checkpoint_written,
                                index=False, encoding="utf-8-sig")
                    _checkpoint_written = True
                except Exception as _cp_exc:
                    print(f"[collect_signals] checkpoint write failed for {stock_id}: {_cp_exc}",
                          file=sys.stderr)
    return signals_by_stock, signal_frames, load_errors, _quota_event.is_set()


def build_daily_snapshot(
    args: argparse.Namespace, client: FinMindClient, config: StrategyConfig
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    end_date = pd.Timestamp(args.end)
    start_date = (end_date - pd.Timedelta(days=args.lookback_days)).strftime("%Y-%m-%d")
    end_date_text = end_date.strftime("%Y-%m-%d")

    market_raw = fetch_market_index(client, start_date, end_date_text)
    if market_raw.empty:
        raise RuntimeError("Unable to download TAIEX market data from FinMind.")
    market = prepare_market_frame(market_raw, config)

    _all_margin_df = pd.DataFrame()
    try:
        _all_margin_df = fetch_all_margin_data(client, end_date_text, lookback=10)
        if not _all_margin_df.empty:
            print(f"[build_daily_snapshot] 融資資料：{len(_all_margin_df)} 筆")
    except Exception as _me:
        print(f"[build_daily_snapshot] ⚠️ 融資資料抓取失敗（降級）：{_me}", file=sys.stderr)

    _all_revenue_df = pd.DataFrame()
    try:
        _all_revenue_df = fetch_all_monthly_revenue(client, end_date_text, lookback_months=13)
    except Exception as _re:
        print(f"[build_daily_snapshot] ⚠️ 月營收資料抓取失敗（降級）：{_re}", file=sys.stderr)

    universe = load_universe(args, client)
    signals_by_stock, _, load_errors, quota_exhausted = collect_signals(
        universe, client, market, config, start_date, end_date_text, args.workers,
        checkpoint_path=getattr(args, "checkpoint_csv", None),
        margin_df=_all_margin_df if not _all_margin_df.empty else None,
        revenue_df=_all_revenue_df if not _all_revenue_df.empty else None,
    )
    snapshot = latest_signal_snapshot(signals_by_stock)
    breadth = compute_market_breadth(snapshot)
    if load_errors:
        breadth["load_errors"] = load_errors[:3]
    breadth["quota_exhausted"] = quota_exhausted
    breadth["market_regime"] = compute_market_regime(market)
    breadth["_snapshot"] = snapshot  # full per-stock data; used by sequential-scan to save all rows
    candidates, watchlist = rank_candidates(
        snapshot,
        args.top_n,
        max_price=args.max_price,
        prefer_lower_price=args.prefer_lower_price,
    )
    # Use the actual latest date present in the data, not the requested end date.
    # FinMind may not have today's data yet when the scan runs intraday.
    actual_date = end_date_text
    if not snapshot.empty and "date" in snapshot.columns:
        try:
            actual_date = pd.Timestamp(snapshot["date"].max()).strftime("%Y-%m-%d")
        except Exception:
            pass
    breadth["actual_date"] = actual_date
    if "stock_id" in snapshot.columns:
        breadth["scanned_ids"] = set(snapshot["stock_id"].astype(str))
    return candidates, watchlist, universe, breadth


_REGIME_EMOJI: dict[str, str] = {"牛市": "🐂", "盤整": "🦀", "熊市": "🐻", "未知": "❓"}


def _format_breadth_line(breadth: dict[str, object]) -> str:
    total = int(breadth.get("total_stocks", 0))
    if not total:
        return "⚠️ **掃描未載入任何股票資料** — 請確認 FINMIND_TOKEN 是否有效，或 FinMind API 是否正常。"
    entry_pct = int(breadth.get("entry_signal_pct", 0))
    above_ema = int(breadth.get("above_ema60", 0))
    trend_up = int(breadth.get("ema60_gt_ema120", 0))
    market_ok = int(breadth.get("market_above_ma60", 0))
    macd_cross = int(breadth.get("macd_golden_cross", 0))
    hist_pos = int(breadth.get("hist_turn_positive", 0))
    mfi_pct = int(breadth.get("mfi_strong", 0))
    ichi_pct = int(breadth.get("above_ichimoku_cloud", 0))
    regime = str(breadth.get("market_regime") or "未知")
    regime_emoji = _REGIME_EMOJI.get(regime, "❓")
    market_tag = "✅" if market_ok > 50 else "⚠️"
    line = (
        f"📊 市場廣度（{total}支）{regime_emoji}`{regime}`｜全條件候選 `{entry_pct}%` | "
        f"站EMA60 `{above_ema}%` | 趨勢 `{trend_up}%` | "
        f"MACD交叉 `{macd_cross}%` | 柱翻正 `{hist_pos}%` | MFI `{mfi_pct}%` | 雲上 `{ichi_pct}%` | 大盤 `{market_tag}`"
    )
    if regime == "熊市" or (market_ok <= 50 and above_ema < 30):
        line += "\n⚠️ **市場偏弱，候選訊號謹慎看待，注意風險控管**"
    return line


def format_scan_message_rich(
    candidates: pd.DataFrame,
    watchlist: pd.DataFrame,
    latest_date: str,
    news_map: dict[str, dict[str, object]] | None = None,
    breadth: dict[str, object] | None = None,
    ai_prediction: dict | None = None,
) -> str:
    news_map = news_map or {}
    lines = [f"🔍 **Taiwan MACD Scan** · {latest_date}", ""]
    if breadth:
        breadth_line = _format_breadth_line(breadth)
        if breadth_line:
            lines.append(breadth_line)
            lines.append("")
    if ai_prediction:
        pred_line = format_prediction_block(ai_prediction, breadth=breadth)
        if pred_line:
            lines.append(pred_line)
            lines.append("")

    # 動能排行榜 — top 5 by momentum_score across candidates + watchlist
    if "momentum_score" in candidates.columns or "momentum_score" in watchlist.columns:
        pool = pd.concat([candidates, watchlist], ignore_index=True)
        if "momentum_score" in pool.columns:
            top_momentum = pool.dropna(subset=["momentum_score"]).sort_values("momentum_score", ascending=False).head(5)
            if not top_momentum.empty:
                lines.append("🔥 **今日動能排行**")
                for rank, (_, row) in enumerate(top_momentum.iterrows(), 1):
                    score = int(row["momentum_score"])
                    bar = "🟢" if score >= 70 else ("🟡" if score >= 40 else "🔴")
                    is_candidate = not candidates.empty and str(row["stock_id"]) in candidates["stock_id"].astype(str).values
                    star = " ✅" if is_candidate else ""
                    lines.append(f"{rank}. **{row['stock_id']}** {row['name']} ｜ 動能 `{score}` {bar}{star}")
                lines.append("")


    if candidates.empty:
        lines.append("今日無全條件候選。")
        # Show which hard condition is blocking the most stocks
        if breadth:
            hard_keys = {
                "macd_golden_cross": "MACD交叉",
                "above_ema60": "站EMA60",
                "ema60_gt_ema120": "EMA60>120",
                "volume_break": "量能放大",
                "rsi_strong": "RSI偏強",
                "breakout_20d": "突破20日高",
                "market_above_ma60": "大盤站MA60",
                "avoid_chase": "未追價",
                "liquidity_ok": "流動性",
            }
            failing = sorted(
                [(label, int(breadth.get(k, 0))) for k, label in hard_keys.items()],
                key=lambda x: x[1],
            )
            if failing:
                bottleneck_parts = [f"`{label}` {pct}%" for label, pct in failing[:4]]
                lines.append(f"🚧 硬條件通過率最低：{' | '.join(bottleneck_parts)}")
        lines.append("")
        # "Only one step away" — stocks missing exactly 1 hard condition
        closest = [
            row for _, row in watchlist.iterrows()
            if not row.get("skip_reason") and _missing_hard_count(row.get("entry_reason")) == 1
        ]
        if closest:
            lines.append("**⏳ 只差一步**")
            for row in closest[:5]:
                missing_label = _missing_hard_labels(row.get("entry_reason"), max_items=1)
                lines.append(
                    f"• **{row['stock_id']}** {row['name']} | 缺 `{missing_label}` | "
                    f"`{_int_or(row['condition_count'])}/{_MAX_CONDITION_COUNT}` | 收 `{_float_or(row['close']):.2f}`"
                )
            lines.append("")
        lines.append("**近似觀察名單**")
        for _, row in watchlist.head(8).iterrows():
            lines.append(_watchlist_line(row, news_map))
        return "\n".join(lines)

    # Sector concentration warning
    if "industry_category" in candidates.columns and len(candidates) >= 3:
        top_sector = candidates["industry_category"].mode()
        if not top_sector.empty:
            top_count = int((candidates["industry_category"] == top_sector.iloc[0]).sum())
            if top_count >= max(3, len(candidates) // 2):
                lines.append(f"⚠️ 候選集中於 **{top_sector.iloc[0]}** ({top_count}/{len(candidates)})，注意分散")
                lines.append("")

    lines.append("**每日候選**")
    for _, row in candidates.head(6).iterrows():
        close = float(row["close"])
        atr = float(row["atr14"]) if pd.notna(row.get("atr14")) else None
        entry_zone, stop, target, rr = _entry_stop_target(close, atr)
        confidence = _confidence_score(row)
        cond_count = int(row["condition_count"])
        price_tag = _low_price_tag(row.get("close"))
        price_note = f" `{price_tag}`" if price_tag else ""
        brief = _news_brief(str(row["stock_id"]), news_map)
        stoch_k = float(row["stoch_k"]) if pd.notna(row.get("stoch_k")) else None
        kd_txt = f" | KD `{stoch_k:.0f}`" if stoch_k is not None else ""
        mfi_val = float(row["mfi14"]) if pd.notna(row.get("mfi14")) else None
        mfi_txt = f" | MFI `{mfi_val:.0f}`" if mfi_val is not None else ""
        invest_streak = _int_or(row.get("invest_trust_streak"))
        invest_txt = f" | 投信 `{invest_streak}d`" if invest_streak >= 1 else ""
        dealer_streak = _int_or(row.get("dealer_buy_streak"))
        dealer_txt = f" | 自營 `{dealer_streak}d`" if dealer_streak >= 1 else ""
        ichi_txt = " | ☁雲上" if bool(row.get("above_ichimoku_cloud")) else ""
        star = " ⭐" if confidence >= 70 else ""
        ret5d = _float_or(row.get("return_5d"))
        ret5d_txt = f" | 5d `{ret5d*100:+.1f}%`" if ret5d != 0 else ""
        industry = str(row.get("industry_category") or "")
        industry_txt = f" 〔{industry}〕" if industry else ""
        swing_low = row.get("close_10d_low")
        swing_txt = f" | 支撐 `{float(swing_low):.2f}`" if swing_low and pd.notna(swing_low) else ""
        trend_block = _trend_block(row)
        momentum = int(row["momentum_score"]) if pd.notna(row.get("momentum_score")) else None
        mom_txt = f" | 動能 `{momentum}`" if momentum is not None else ""
        lines.append(f"━━━━━━━━━━━━━━━━━━━━")
        lines.append(f"**{row['stock_id']}** {row['name']}{industry_txt}{price_note}{star}  信心 `{confidence}/100` | `{cond_count}/{_MAX_CONDITION_COUNT}條`{mom_txt}")
        lines.append(f"💰 收 `{close:.2f}` | 進場 `{entry_zone}` | 停損 `{stop}` | 目標 `{target}` | R:R `{rr}`{swing_txt}")
        lines.append(f"📊 RSI `{_float_or(row.get('rsi14')):.1f}` | ADX `{_float_or(row.get('adx14')):.1f}`{kd_txt}{mfi_txt} | 量比 `{_float_or(row.get('volume_ratio')):.1f}x`{ret5d_txt}{ichi_txt}")
        lines.append(f"🏦 外資連買 `{_int_or(row.get('foreign_buy_streak'))}d`{invest_txt}{dealer_txt}")
        if trend_block:
            lines.append(trend_block)
        lines.append(f"🔍 {_reason_labels(row.get('entry_reason'), max_items=5)}")
        if brief:
            lines.append(f"  {brief.strip()}")
    if not watchlist.empty:
        lines.extend(["", "━━━━━━━━━━━━━━━━━━━━", "**近似觀察名單**"])
        for _, row in watchlist.head(5).iterrows():
            lines.append(_watchlist_line(row, news_map))
    return "\n".join(lines)


def format_hybrid_message_rich(
    candidates: pd.DataFrame,
    watchlist: pd.DataFrame,
    live_quotes: pd.DataFrame,
    latest_date: str,
    news_map: dict[str, dict[str, object]] | None = None,
) -> str:
    news_map = news_map or {}
    lines = [f"📊 **Taiwan Hybrid Monitor** · {latest_date}", "", "**Step 1｜MACD 每日篩選**"]
    if candidates.empty:
        lines.append("今日無全條件候選。")
    else:
        for _, row in candidates.head(5).iterrows():
            price_tag = _low_price_tag(row.get("close"))
            price_note = f" `{price_tag}`" if price_tag else ""
            brief = _news_brief(str(row["stock_id"]), news_map)
            lines.append(
                f"• **{row['stock_id']}** {row['name']} | 收 `{float(row['close']):.2f}`{price_note} | {_reason_labels(row.get('entry_reason'))}{brief}"
            )
    if not watchlist.empty:
        lines.extend(["", "**Step 2｜近似觀察名單**"])
        for _, row in watchlist.head(5).iterrows():
            close_value = row.get("close") if pd.notna(row.get("close")) else None
            close_text = f"`{float(close_value):.2f}`" if close_value is not None else "N/A"
            price_tag = _low_price_tag(close_value)
            price_note = f" `{price_tag}`" if price_tag else ""
            score_value = row.get("condition_count") if pd.notna(row.get("condition_count")) else None
            score_text = f"`{int(score_value)}/23`" if score_value is not None else "manual"
            brief = _news_brief(str(row["stock_id"]), news_map)
            lines.append(
                f"• **{row['stock_id']}** {row['name']} | 收 {close_text}{price_note} | {score_text} | {_reason_labels(row.get('entry_reason'))}{brief}"
            )
    if not live_quotes.empty:
        lines.extend(["", "**Step 3｜即時報價**"])
        for _, row in live_quotes.iterrows():
            if pd.notna(row.get("error")):
                lines.append(f"• {row['symbol']} — 報價錯誤: {row['error']}")
                continue
            intraday = row.get("intraday_pct")
            intraday_text = f"`{float(intraday) * 100:.2f}%`" if pd.notna(intraday) else "N/A"
            lines.append(
                f"• **{row['symbol']}** {row.get('name') or ''} | 最新 `{row.get('last')}` | 高 `{row.get('high')}` | 量 `{row.get('volume')}` | 漲幅 {intraday_text}"
            )
    return "\n".join(lines)


def detect_quote_events(
    current_quotes: pd.DataFrame,
    previous_state: dict[str, dict[str, float | str]],
    rise_threshold: float,
    drop_threshold: float,
    volume_multiplier: float,
    daily_volume_lookup: dict[str, float] | None = None,
) -> tuple[list[dict[str, object]], dict[str, dict[str, float | str]]]:
    events: list[dict[str, object]] = []
    next_state = dict(previous_state)
    for _, row in current_quotes.iterrows():
        symbol = str(row.get("symbol") or "")
        if not symbol or pd.notna(row.get("error")):
            continue
        last_value = pd.to_numeric(row.get("last"), errors="coerce")
        high_value = pd.to_numeric(row.get("high"), errors="coerce")
        volume_value = pd.to_numeric(row.get("volume"), errors="coerce")
        intraday_value = pd.to_numeric(row.get("intraday_pct"), errors="coerce")
        previous = next_state.get(symbol, {})
        previous_high = float(previous.get("high") or 0)
        previous_intraday = float(previous.get("intraday_pct") or 0)
        previous_volume = float(previous.get("volume") or 0)

        row_dict: dict[str, object] | None = None  # lazy — only convert if an event fires

        if (pd.notna(high_value) and high_value > previous_high
                and pd.notna(last_value) and last_value >= high_value * 0.998
                and pd.notna(intraday_value) and intraday_value >= 0.015):
            row_dict = row_dict or row.to_dict()
            events.append({"symbol": symbol, "event_key": "breakout_high", "label": "突破盤中新高", "row": row_dict})
        if pd.notna(intraday_value) and intraday_value >= rise_threshold and previous_intraday < rise_threshold:
            row_dict = row_dict or row.to_dict()
            events.append({"symbol": symbol, "event_key": "sharp_rise", "label": f"急拉 {intraday_value * 100:.2f}%", "row": row_dict})
        if pd.notna(intraday_value) and intraday_value <= drop_threshold and previous_intraday > drop_threshold:
            row_dict = row_dict or row.to_dict()
            events.append({"symbol": symbol, "event_key": "sharp_drop", "label": f"急跌 {intraday_value * 100:.2f}%", "row": row_dict})

        daily_baseline = (daily_volume_lookup or {}).get(symbol, 0)
        if daily_baseline > 0:
            if pd.notna(volume_value) and volume_value >= daily_baseline * volume_multiplier:
                ratio = volume_value / daily_baseline
                row_dict = row_dict or row.to_dict()
                events.append({"symbol": symbol, "event_key": "volume_surge", "label": f"量能放大 x{ratio:.1f}（vs 日均）", "row": row_dict})
        elif pd.notna(volume_value) and previous_volume > 0 and volume_value >= previous_volume * volume_multiplier:
            row_dict = row_dict or row.to_dict()
            events.append({"symbol": symbol, "event_key": "volume_surge", "label": f"量能放大 x{volume_value / previous_volume:.1f}", "row": row_dict})

        next_state[symbol] = {
            "high": float(high_value) if pd.notna(high_value) else previous_high,
            "intraday_pct": float(intraday_value) if pd.notna(intraday_value) else previous_intraday,
            "volume": float(volume_value) if pd.notna(volume_value) else previous_volume,
            "quote_time": str(row.get("quote_time") or ""),
        }
    return events, next_state


_EVENT_EMOJI: dict[str, str] = {
    "breakout_high": "🆙",
    "sharp_rise": "📈",
    "sharp_drop": "📉",
    "volume_surge": "🔥",
}


def format_event_message(
    event: dict[str, object],
    snapshot_lookup: dict[str, dict[str, object]],
    news_map: dict[str, dict[str, object]] | None = None,
) -> str:
    news_map = news_map or {}
    row = event["row"]
    symbol = str(event["symbol"])
    snapshot = snapshot_lookup.get(symbol, {})
    name = row.get("name") or snapshot.get("name") or ""
    last_value = row.get("last")
    high_value = row.get("high")
    volume_value = row.get("volume")
    intraday_value = row.get("intraday_pct")
    intraday_text = f"**{float(intraday_value) * 100:.2f}%**" if pd.notna(intraday_value) else "N/A"
    price_tag = _low_price_tag(last_value)
    price_note = f" `{price_tag}`" if price_tag else ""
    reason_text = _reason_labels(snapshot.get("entry_reason"))
    emoji = _EVENT_EMOJI.get(str(event["event_key"]), "⚠️")
    news_text = _news_brief(symbol, news_map).strip()
    confidence = _confidence_score(snapshot) if snapshot else 0
    star = " ⭐" if confidence >= 70 else ""
    rsi_txt = f" RSI `{float(snapshot['rsi14']):.0f}`" if snapshot.get("rsi14") and pd.notna(snapshot.get("rsi14")) else ""
    mfi_txt = f" MFI `{float(snapshot['mfi14']):.0f}`" if snapshot.get("mfi14") and pd.notna(snapshot.get("mfi14")) else ""
    # ATR-based stop/target when ATR data is available from the daily snapshot
    risk_line = ""
    try:
        last_f = float(last_value)  # type: ignore[arg-type]
        atr_val = float(snapshot.get("atr14") or 0) or None
        _, stop, target, rr = _entry_stop_target(last_f, atr_val)
        risk_line = f"停損 `{stop}` | 目標 `{target}` | R:R `{rr}`"
    except (TypeError, ValueError):
        pass
    parts = [
        f"{emoji} **即時警報｜{event['label']}**",
        f"**{symbol}** {name}{price_note}{star}  信心 `{confidence}/100`",
        f"最新 `{last_value}` | 高點 `{high_value}` | 量 `{volume_value}` | 漲幅 {intraday_text}",
    ]
    if risk_line:
        parts.append(risk_line)
    indicator_line = f"技術: {reason_text}"
    if rsi_txt or mfi_txt:
        indicator_line += f" |{rsi_txt}{mfi_txt}"
    parts.append(indicator_line)
    parts.append(f"新聞: {news_text if news_text else '無近期消息'}")
    return "\n".join(parts)


# ── Volume-surge detection (rolling incremental volume vs N-minute average) ────

_VOL_WINDOW = 5       # rolling average window (samples; 1 sample ≈ 1 poll interval)
_VOL_MIN_SAMPLES = 3  # need this many past samples before alerting


def detect_volume_surges(
    current_quotes: pd.DataFrame,
    prev_cumvol: dict[str, float],
    vol_history: dict[str, "deque[float]"],
    multiplier: float,
) -> tuple[list[dict[str, object]], dict[str, float]]:
    """Compare each interval's incremental volume against the rolling average.

    Fugle returns cumulative daily volume, so we diff successive readings to
    get per-interval volume, then compare against the rolling mean of recent
    intervals.  No alert on the first observation (no baseline yet).
    """
    events: list[dict[str, object]] = []
    next_cumvol = dict(prev_cumvol)

    for _, row in current_quotes.iterrows():
        symbol = str(row.get("symbol") or "")
        if not symbol or pd.notna(row.get("error")):
            continue
        raw_vol = pd.to_numeric(row.get("volume"), errors="coerce")
        if pd.isna(raw_vol):
            continue

        cumvol = float(raw_vol)
        hist = vol_history.setdefault(symbol, deque(maxlen=_VOL_WINDOW))

        if symbol not in prev_cumvol:
            next_cumvol[symbol] = cumvol
            continue

        incr = max(0.0, cumvol - prev_cumvol[symbol])
        next_cumvol[symbol] = cumvol

        if len(hist) >= _VOL_MIN_SAMPLES and incr > 0:
            avg = sum(hist) / len(hist)
            if avg > 0 and incr >= avg * multiplier:
                ratio = incr / avg
                events.append({
                    "symbol": symbol,
                    "label": f"爆量 ×{ratio:.1f}（近{len(hist)}分均 {avg:,.0f}）",
                    "incr_vol": incr,
                    "avg_vol": avg,
                    "row": row.to_dict(),
                })

        if incr > 0:
            hist.append(incr)

    return events, next_cumvol


def format_volume_alert(
    event: dict[str, object],
    snapshot_lookup: dict[str, dict[str, object]],
) -> str:
    row = event["row"]
    symbol = str(event["symbol"])
    snapshot = snapshot_lookup.get(symbol, {})
    name = row.get("name") or snapshot.get("name") or ""
    last = row.get("last")
    incr = float(event.get("incr_vol", 0))
    avg  = float(event.get("avg_vol", 0))
    intraday = pd.to_numeric(row.get("intraday_pct"), errors="coerce")
    pct_txt = f"{float(intraday) * 100:+.2f}%" if pd.notna(intraday) else "N/A"
    rsi_txt = (
        f" | RSI `{float(snapshot['rsi14']):.0f}`"
        if snapshot.get("rsi14") and pd.notna(snapshot.get("rsi14")) else ""
    )
    return "\n".join([
        f"🔥 **爆量警報** · {_cst_now()} CST",
        f"**{symbol}** {name}",
        f"最新 `{last}` | 漲幅 `{pct_txt}`{rsi_txt}",
        event["label"],
    ])


def format_heartbeat_message(quotes: pd.DataFrame, date: str) -> str:
    now = _cst_now()
    lines = [f"💓 **盤中快報** · {now} CST ({date})", ""]
    if quotes.empty:
        lines.append("無即時報價。")
        return "\n".join(lines)
    for _, row in quotes.iterrows():
        if pd.notna(row.get("error")):
            lines.append(f"• **{row['symbol']}** — 報價錯誤")
            continue
        intraday = row.get("intraday_pct")
        if pd.notna(intraday):
            pct = float(intraday) * 100
            arrow = "📈" if pct > 0.5 else ("📉" if pct < -0.5 else "➡️")
            pct_text = f"`{pct:+.2f}%`"
        else:
            arrow, pct_text = "➡️", "N/A"
        vol = row.get("volume")
        vol_part = f" | 量 `{vol}`" if vol is not None and str(vol) not in ("None", "nan", "") else ""
        lines.append(
            f"• **{row['symbol']}** {row.get('name') or ''} | {arrow} {pct_text} | 收 `{row.get('last')}`{vol_part}"
        )
    return "\n".join(lines)


def format_daily_report_message(
    candidates: pd.DataFrame,
    watchlist: pd.DataFrame,
    closing_quotes: pd.DataFrame,
    date: str,
    news_map: dict[str, dict[str, object]] | None = None,
    breadth: dict[str, object] | None = None,
) -> str:
    news_map = news_map or {}
    lines = [f"📋 **盤後總結** · {date}", ""]
    if breadth:
        breadth_line = _format_breadth_line(breadth)
        if breadth_line:
            lines.append(breadth_line)
            lines.append("")
    lines.append("**今日候選（篩選時收盤）**")
    if candidates.empty:
        lines.append("今日無全條件候選。")
    else:
        for _, row in candidates.head(8).iterrows():
            price_tag = _low_price_tag(row.get("close"))
            price_note = f" `{price_tag}`" if price_tag else ""
            brief = _news_brief(str(row["stock_id"]), news_map)
            lines.append(
                f"• **{row['stock_id']}** {row['name']} | 收 `{row['close']:.2f}`{price_note} | {_reason_labels(row.get('entry_reason'))}{brief}"
            )
    if not closing_quotes.empty:
        ok_rows = closing_quotes[closing_quotes["error"].isna()] if "error" in closing_quotes.columns else closing_quotes
        err_rows = closing_quotes[closing_quotes["error"].notna()] if "error" in closing_quotes.columns else pd.DataFrame()
        if ok_rows.empty and not err_rows.empty:
            sample_err = str(err_rows["error"].iloc[0]) if not err_rows.empty else ""
            if "not configured" in sample_err or not sample_err:
                err_hint = "FUGLE_API_KEY 未設定，請至 GitHub Secrets 加入金鑰"
            elif "401" in sample_err or "金鑰" in sample_err:
                err_hint = "金鑰無效或未授權，請確認 FUGLE_API_KEY 是否正確"
            elif "403" in sample_err:
                err_hint = "金鑰無此 API 權限，請升級 Fugle 方案"
            elif "429" in sample_err or "頻率" in sample_err:
                err_hint = "請求頻率過高，稍後重試"
            else:
                err_hint = sample_err[:60]
            lines.extend(["", "**盤後報價對比**", f"⚠️ Fugle API 無法取得報價（{err_hint}）"])
        elif not ok_rows.empty:
            lines.extend(["", "**盤後報價對比**"])
            for _, row in ok_rows.iterrows():
                intraday = row.get("intraday_pct")
                if pd.notna(intraday):
                    pct = float(intraday) * 100
                    arrow = "📈" if pct > 1 else ("📉" if pct < -1 else "➡️")
                    pct_text = f"**{pct:+.2f}%**"
                else:
                    arrow, pct_text = "➡️", "N/A"
                lines.append(
                    f"• **{row['symbol']}** {row.get('name') or ''} | {arrow} {pct_text} | 收 `{row.get('last')}` | 量 `{row.get('volume')}`"
                )
    if not watchlist.empty:
        lines.extend(["", "**近似名單（今日未達全條件）**"])
        for _, row in watchlist.head(5).iterrows():
            lines.append(_watchlist_line(row, news_map))
    return "\n".join(lines)


def collect_intraday_snapshot(
    client: FinMindClient,
    watch_symbols: list[str],
    trade_date: str,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for symbol in watch_symbols:
        try:
            kbars = fetch_stock_kbar(client, symbol, trade_date)
            if kbars.empty:
                rows.append({"stock_id": symbol, "error": "No KBar data returned"})
                continue
            latest = kbars.iloc[-1].to_dict()
            rows.append(latest)
        except Exception as error:
            rows.append({"stock_id": symbol, "error": str(error)})
    return pd.DataFrame(rows)


def format_sponsor_message(
    candidates: pd.DataFrame,
    watchlist: pd.DataFrame,
    intraday_rows: pd.DataFrame,
    latest_date: str,
    cycle: int,
    repeat_count: int,
) -> str:
    lines = [f"Taiwan sponsor monitor ({latest_date}) cycle {cycle}/{repeat_count}", "", "Step 1: daily prefilter"]
    if candidates.empty:
        lines.append("No full-match candidates today.")
    else:
        for _, row in candidates.head(5).iterrows():
            close_value = row.get("close") if pd.notna(row.get("close")) else None
            score_value = row.get("condition_count") if pd.notna(row.get("condition_count")) else None
            close_text = f"{float(close_value):.2f}" if close_value is not None else "N/A"
            score_text = f"{int(score_value)}/23" if score_value is not None else "manual"
            lines.append(f"- {row['stock_id']} {row['name']} | close {close_text} | score {score_text}")

    if not watchlist.empty:
        lines.extend(["", "Step 2: tracked symbols"])
        for _, row in watchlist.head(5).iterrows():
            close_value = row.get("close") if pd.notna(row.get("close")) else None
            score_value = row.get("condition_count") if pd.notna(row.get("condition_count")) else None
            close_text = f"{float(close_value):.2f}" if close_value is not None else "N/A"
            score_text = f"{int(score_value)}/23" if score_value is not None else "manual"
            lines.append(f"- {row['stock_id']} {row['name']} | close {close_text} | score {score_text}")

    if not intraday_rows.empty:
        lines.extend(["", "Step 3: intraday KBar snapshot"])
        for _, row in intraday_rows.iterrows():
            if pd.notna(row.get("error")):
                lines.append(f"- {row.get('stock_id')} | KBar error: {row['error']}")
                continue
            lines.append(
                f"- {row.get('stock_id')} | {row.get('minute')} | O {row.get('open')} H {row.get('high')} L {row.get('low')} C {row.get('close')} V {row.get('volume')}"
            )
    return "\n".join(lines)


def run_scan(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    token_ok, token_msg = validate_finmind_token()
    if not token_ok:
        err = f"❌ **掃描中止** · {args.end}\n{token_msg}"
        _safe_print(err)
        if args.notify:
            send_discord_messages([err])
        return

    _batch_idx = getattr(args, "batch_index", -1)
    _batch_count = getattr(args, "batch_count", 15)
    already_scanned: set[str] = set()
    total_universe = 0
    new_to_scan: set[str] = set()

    # In batch mode, probe quota before launching a scan
    if _batch_idx >= 0:
        quota_ok, quota_msg = probe_batch_quota(client)
        if not quota_ok:
            _safe_print(f"[batch {_batch_idx}] {quota_msg}")
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([
                    f"⏰ **批次 {_batch_idx}/{_batch_count - 1} 跳過** · 配額不足 · {_cst_now()} CST\n{quota_msg}"
                ])
            return

    # In batch mode, detect already-scanned stocks and only scan new ones
    if _batch_idx >= 0:
        import copy as _copy
        _scan_dir = Path(args.output) / "full_scan"
        _scan_dir.mkdir(parents=True, exist_ok=True)

        for _p in sorted(_scan_dir.glob("batch_*.csv")):
            try:
                _df = pd.read_csv(_p, encoding="utf-8-sig")
                if "stock_id" in _df.columns:
                    already_scanned.update(_df["stock_id"].astype(str))
            except Exception:
                pass

        _stock_info = fetch_stock_info(client)
        _full_univ = build_auto_universe(_stock_info, max_symbols=args.max_universe)
        total_universe = len(_full_univ)
        _bc = max(1, _batch_count)
        _size = (len(_full_univ) + _bc - 1) // _bc
        _start = _batch_idx * _size
        _batch_slice = _full_univ.iloc[_start: _start + _size]
        _batch_ids = set(_batch_slice["stock_id"].astype(str))
        new_to_scan = _batch_ids - already_scanned

        _safe_print(
            f"[batch {_batch_idx}] 本批 {len(_batch_ids)} 支：已掃 {len(_batch_ids & already_scanned)}，"
            f"新掃 {len(new_to_scan)}，全局累計 {len(already_scanned)}/{total_universe}"
        )

        if not new_to_scan:
            msg = (
                f"✅ **批次 {_batch_idx} 全部已掃** · {_cst_now()} CST\n"
                f"全局累計 `{len(already_scanned)}/{total_universe}` 支"
            )
            _safe_print(msg)
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([msg])
            return

        _gap_file = _scan_dir / f"_scan_new_{_batch_idx:02d}.csv"
        _new_univ = _batch_slice[_batch_slice["stock_id"].astype(str).isin(new_to_scan)]
        _new_univ.to_csv(_gap_file, index=False)
        args = _copy.copy(args)
        args.stocks = str(_gap_file)
        args.batch_index = -1

    if args.notify:
        send_discord_messages([f"🔄 **選股掃描開始** · {_cst_now()} CST · {args.end}"])
    candidates, watchlist, universe, breadth = build_daily_snapshot(args, client, config)
    latest_date = str(breadth.get("actual_date") or args.end)

    # If 0 stocks loaded despite a valid token, diagnose and report why
    if breadth.get("total_stocks", 0) == 0:
        load_errors: list[str] = list(breadth.get("load_errors") or [])
        diag_lines = [f"⚠️ **診斷報告** · {latest_date}"]
        diag_lines.append("• Token 驗證：✅ 通過")
        diag_lines.append(f"• 股票清單（universe）：{len(universe)} 支")
        if len(universe) == 0:
            diag_lines.append("  → build_auto_universe 回傳空清單，請確認 TaiwanStockInfo 可用")
        elif load_errors:
            diag_lines.append("• 抓取失敗樣本（前 3 筆）：")
            for e in load_errors:
                diag_lines.append(f"  → `{e}`")
        else:
            diag_lines.append(f"  → {len(universe)} 支股票全部回傳空資料（無例外）")
            diag_lines.append("  → TaiwanStockPrice 可能對此帳戶返回空結果，請查看 Actions logs")
        diag_msg = "\n".join(diag_lines)
        _safe_print(diag_msg)
        if args.notify:
            send_discord_messages([diag_msg])

    # Apply minimum confidence filter for notifications (vectorized)
    min_conf = getattr(args, "min_confidence", 0)
    notify_candidates = candidates
    if min_conf > 0 and not candidates.empty:
        cond_v = pd.to_numeric(candidates.get("condition_count", 0), errors="coerce").fillna(0)
        adx_v = pd.to_numeric(candidates.get("adx14", 0), errors="coerce").fillna(0)
        rs_v = pd.to_numeric(candidates.get("relative_strength_5d", 0), errors="coerce").fillna(0)
        vol_v = pd.to_numeric(candidates.get("volume_ratio", 0), errors="coerce").fillna(0)
        scores = (
            (cond_v / 23 * 55).clip(0, 55)
            + (adx_v / 40 * 20).clip(0, 20)
            + (rs_v * 200).clip(0, 15)
            + ((vol_v - 1) / 2 * 10).clip(0, 10)
        ).astype(int).clip(0, 100)
        notify_candidates = candidates[scores >= min_conf].copy()

    # AI market-direction prediction (best-effort; silent if deps missing)
    ai_prediction: dict | None = None
    try:
        train_start = (pd.Timestamp(args.end) - pd.Timedelta(days=365)).strftime("%Y-%m-%d")
        market_train = fetch_market_index(client, train_start, args.end)
        if not market_train.empty:
            us_df = fetch_us_features(train_start, args.end)
            predictor = MarketPredictor(horizon=5)
            predictor.fit(market_train, us_df if not us_df.empty else None)
            ai_prediction = predictor.predict_proba(market_train, us_df if not us_df.empty else None)
    except Exception as _pred_exc:
        _safe_print(f"[ai] 預測略過：{_pred_exc}")

    # In batch mode, merge new results with existing batch CSV and save
    if _batch_idx >= 0:
        _save_dir = Path(args.output) / "full_scan"
        _save_dir.mkdir(parents=True, exist_ok=True)
        _batch_csv = _save_dir / f"batch_{_batch_idx:02d}.csv"
        _batch_df = pd.concat([candidates, watchlist], ignore_index=True) if not watchlist.empty else candidates
        _batch_df = _batch_df.sort_values("entry_score", ascending=False).drop_duplicates(subset=["stock_id"])

        if _batch_csv.exists():
            _existing_df = pd.read_csv(_batch_csv, encoding="utf-8-sig")
            _batch_df = pd.concat([_existing_df, _batch_df], ignore_index=True)
            _batch_df = _batch_df.sort_values("entry_score", ascending=False).drop_duplicates(subset=["stock_id"])

        _batch_df.to_csv(_batch_csv, index=False, encoding="utf-8-sig")
        _batch_df.to_excel(_batch_csv.with_suffix(".xlsx"), index=False, engine="openpyxl")
        _safe_print(f"[batch {_batch_idx}] 儲存候選+觀察：{len(_batch_df)} 檔（候選 {len(candidates)}，觀察 {len(watchlist)}）→ {_batch_csv}")
        _write_batch_markdown(
            _save_dir, _batch_idx, _batch_count,
            args.end, candidates, watchlist, breadth,
        )
        _new_global = len(already_scanned) + len(new_to_scan)
        if os.getenv("DISCORD_WEBHOOK_URL"):
            _cand_n = len(candidates)
            _watch_n = len(watchlist)
            _status = "✅" if _cand_n > 0 else "🔵"
            send_discord_messages([
                f"{_status} **批次 {_batch_idx} 完成** · {_cst_now()} CST · {args.end}\n"
                f"新掃 `{breadth.get('total_stocks', 0)}` 支 | 候選 `{_cand_n}` | 觀察 `{_watch_n}`\n"
                f"全局累計 `{_new_global}/{total_universe}` 支"
            ])
        (_save_dir / f"_scan_new_{_batch_idx:02d}.csv").unlink(missing_ok=True)

    report_path = save_scan_report(args.output, candidates, watchlist, universe)
    news_map = {}
    if args.include_news:
        scan_rows = notify_candidates.to_dict("records") if not notify_candidates.empty else watchlist.to_dict("records")
        news_map = build_news_map(args.output, scan_rows, news_limit=args.news_limit)
    message = format_scan_message_rich(notify_candidates, watchlist, latest_date, news_map=news_map, breadth=breadth, ai_prediction=ai_prediction)
    _safe_print(message)
    _safe_print("")
    _safe_print(f"Scan report: {report_path}")
    if args.notify:
        send_discord_messages(split_message(message))
    if notion_enabled():
        sync_scan_results(candidates, watchlist, latest_date, news_map, market_regime=str(breadth.get("market_regime") or ""))


def run_fill_batch_gaps(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    """批次掃完後，比對 universe 該批切片，補掃缺漏股票，合併回同批 CSV/XLSX。"""
    import copy
    batch_dir = Path(args.output) / "full_scan"
    batch_idx = getattr(args, "batch_index", -1)
    if batch_idx < 0:
        _safe_print("[fill-batch-gaps] 未指定 --batch-index，略過")
        return
    batch_csv = batch_dir / f"batch_{batch_idx:02d}.csv"
    if not batch_csv.exists():
        _safe_print(f"[fill-batch-gaps] {batch_csv} 不存在，略過")
        return

    existing = pd.read_csv(batch_csv, encoding="utf-8-sig")
    scanned = set(existing["stock_id"].astype(str)) if "stock_id" in existing.columns else set()

    stock_info = fetch_stock_info(client)
    full_univ = build_auto_universe(stock_info, max_symbols=args.max_universe)
    bc = max(1, getattr(args, "batch_count", 15))
    n = len(full_univ)
    size = (n + bc - 1) // bc
    start = batch_idx * size
    batch_univ = full_univ.iloc[start: start + size]
    missing = set(batch_univ["stock_id"].astype(str)) - scanned

    if not missing:
        _safe_print(f"[fill-batch-gaps] 批次 {batch_idx} 無缺漏")
        return
    _safe_print(f"[fill-batch-gaps] 批次 {batch_idx} 缺漏 {len(missing)} 支，開始補掃")

    quota_ok, quota_msg = probe_batch_quota(client)
    if not quota_ok:
        _safe_print(f"[fill-batch-gaps] 配額不足，略過：{quota_msg}")
        return

    gap_univ = batch_univ[batch_univ["stock_id"].astype(str).isin(missing)]
    gap_file = batch_dir / f"_gap_{batch_idx:02d}.csv"
    gap_univ.to_csv(gap_file, index=False)

    gap_args = copy.copy(args)
    gap_args.stocks = str(gap_file)
    gap_args.batch_index = -1

    try:
        candidates, watchlist, _, _ = build_daily_snapshot(gap_args, client, config)
        gap_df = pd.concat([candidates, watchlist], ignore_index=True) if not watchlist.empty else candidates
        if not gap_df.empty:
            merged = pd.concat([existing, gap_df], ignore_index=True)
            merged = merged.sort_values("entry_score", ascending=False).drop_duplicates(subset=["stock_id"])
            merged.to_csv(batch_csv, index=False, encoding="utf-8-sig")
            merged.to_excel(batch_csv.with_suffix(".xlsx"), index=False, engine="openpyxl")
            _safe_print(f"[fill-batch-gaps] 補入 {len(gap_df)} 支，合併後共 {len(merged)} 支")
    finally:
        gap_file.unlink(missing_ok=True)


def run_fill_gaps(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    """彙整前：比對所有批次 CSV vs 完整 universe，補掃仍缺漏的股票 → batch_gap.csv/xlsx。

    讀取順序：
      1. _attempted_{date}.csv — sequential-scan 記錄的「所有已嘗試」股票（含無訊號者）
      2. batch_*.csv           — 舊批次掃描結果（相容舊架構）
    兩者合集 = 真正的「已掃過」清單，差集 = 需補掃。
    """
    import copy
    batch_dir = Path(args.output) / "full_scan"
    scanned: set[str] = set()

    # 1. Read sequential-scan's complete attempt log (preferred; covers no-signal stocks too)
    for p in sorted(batch_dir.glob("_attempted_*.csv")):
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if "stock_id" in df.columns:
                scanned.update(df["stock_id"].astype(str))
        except Exception:
            pass

    # 2. Read old-style batch_NN.csv / batch_gap.csv for compatibility with parallel-scan results.
    #    Skip batch_seq*.csv — those are sequential-scan outputs already covered by _attempted_*.csv.
    for p in batch_dir.glob("batch_*.csv"):
        if p.stem.startswith("batch_seq"):
            continue
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if "stock_id" in df.columns:
                scanned.update(df["stock_id"].astype(str))
        except Exception:
            pass

    stock_info = fetch_stock_info(client)
    full_univ = build_auto_universe(stock_info, max_symbols=args.max_universe)
    missing = set(full_univ["stock_id"].astype(str)) - scanned

    if not missing:
        _safe_print("[fill-gaps] 無全局缺漏，跳過")
        return
    _safe_print(f"[fill-gaps] 全局缺漏 {len(missing)} 支，開始補掃")

    gap_univ = full_univ[full_univ["stock_id"].astype(str).isin(missing)]
    gap_file = batch_dir / "_gap_global.csv"
    gap_univ.to_csv(gap_file, index=False)

    gap_args = copy.copy(args)
    gap_args.stocks = str(gap_file)
    gap_args.batch_index = -1

    try:
        candidates, watchlist, _, _ = build_daily_snapshot(gap_args, client, config)
        gap_df = pd.concat([candidates, watchlist], ignore_index=True) if not watchlist.empty else candidates
        gap_df = gap_df.sort_values("entry_score", ascending=False).drop_duplicates(subset=["stock_id"])
        gap_df.to_csv(batch_dir / "batch_gap.csv", index=False, encoding="utf-8-sig")
        gap_df.to_excel(batch_dir / "batch_gap.xlsx", index=False, engine="openpyxl")
        _safe_print(f"[fill-gaps] 補掃完成：{len(gap_df)} 支 → batch_gap.csv")
    finally:
        gap_file.unlink(missing_ok=True)


def run_sequential_scan(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    """
    Hourly sequential multi-token scan with daily accumulation.

    Two file types per day:
      _attempted_{date}.csv      — ALL stock IDs that were fetched (incl. no-signal).
                                   Used for resume tracking across hourly runs.
      batch_seq{N}_{date}.csv/xlsx — raw candidates+watchlist per token, saved immediately
                                     after each token finishes — NO sorting, NO merging.
                                     Aggregate at 20:00 reads all batch_*.csv files and
                                     produces the final sorted ranking + TOP 20.

    Each hourly run:
      1. Reads _attempted_{date}.csv to know which stocks were already attempted
      2. Probes each token in order; scans remaining stocks; auto-switches on quota exhaustion
      3. Appends newly attempted IDs to _attempted_{date}.csv immediately after each token
      4. Dumps raw results to batch_seq{N}_{date}.csv/xlsx — no merging across tokens
    """
    import copy as _copy

    today = args.end
    scan_dir = Path(args.output) / "full_scan"
    scan_dir.mkdir(parents=True, exist_ok=True)
    _seg_idx = getattr(args, "batch_index", -1)
    _seg_suffix = f"_seg{_seg_idx}" if _seg_idx >= 0 else ""
    attempted_csv = scan_dir / f"_attempted_{today}{_seg_suffix}.csv"
    no_data_csv = scan_dir / "_no_data_stocks.csv"  # persistent cross-day blacklist

    # Load which stocks were already attempted today (all fetched, incl. no-signal)
    already_scanned: set[str] = set()
    if attempted_csv.exists():
        try:
            _att = pd.read_csv(attempted_csv, encoding="utf-8-sig")
            if "stock_id" in _att.columns:
                already_scanned = set(_att["stock_id"].astype(str))
        except Exception:
            pass

    # Load persistent blacklist of stocks that returned no price data on any previous day
    # (delisted stocks, stocks with insufficient history, FinMind missing data, etc.)
    no_data_ids: set[str] = set()
    if no_data_csv.exists():
        try:
            _nd = pd.read_csv(no_data_csv, encoding="utf-8-sig")
            if "stock_id" in _nd.columns:
                no_data_ids = set(_nd["stock_id"].astype(str))
        except Exception:
            pass

    # Get full universe — try each token in order so account 0 quota exhaustion doesn't crash.
    # Track tokens confirmed as quota-exhausted so the scan loop skips them immediately.
    _all_env_tokens = [
        ("FINMIND_TOKEN",   os.getenv("FINMIND_TOKEN")),
        ("FINMIND_TOKEN_2", os.getenv("FINMIND_TOKEN_2")),
        ("FINMIND_TOKEN_3", os.getenv("FINMIND_TOKEN_3")),
        ("FINMIND_TOKEN_4", os.getenv("FINMIND_TOKEN_4")),
        ("FINMIND_TOKEN_5", os.getenv("FINMIND_TOKEN_5")),
        ("FINMIND_TOKEN_6", os.getenv("FINMIND_TOKEN_6")),
        ("FINMIND_TOKEN_7", os.getenv("FINMIND_TOKEN_7")),
        ("FINMIND_TOKEN_8", os.getenv("FINMIND_TOKEN_8")),
        ("FINMIND_TOKEN_9", os.getenv("FINMIND_TOKEN_9")),
    ]
    _orig_token_env = os.environ.get("FINMIND_TOKEN", "")
    _quota_exhausted_keys: set[str] = set()  # legacy marker; no hard-skip in sequential mode
    full_univ = None
    try:
        for _env_key, _env_val in _all_env_tokens:
            if not _env_val:
                continue
            try:
                os.environ["FINMIND_TOKEN"] = _env_val
                _info_client = FinMindClient(cache_dir=client.cache_dir)
                stock_info = fetch_stock_info(_info_client)
                full_univ = build_auto_universe(stock_info, max_symbols=args.max_universe)
                if len(full_univ) > 0:
                    _safe_print(f"[sequential] 股票清單取得成功（使用 {_env_key}）：{len(full_univ)} 支")
                    break
            except Exception as _e:
                _exc_s = str(_e).lower()
                # Only mark as quota-exhausted for permanent daily limit errors.
                # Transient rate-limits ("transient rate-limit", "rate limit (402)") must NOT
                # be treated as permanent — the account may still have quota for the scan.
                # Do not hard-mark this token as exhausted here.
                # TaiwanStockInfo probe can be transiently throttled and does not always
                # reflect true per-account scan availability at handoff time.
                _safe_print(f"[sequential] 取股票清單失敗（{_env_key}）：{_e}，嘗試下一個 token")
                if os.getenv("DISCORD_WEBHOOK_URL"):
                    send_discord_messages([
                        f"⚠️ **股票清單取得失敗（{_env_key}）** · {_cst_now()} CST\n"
                        f"改用下一個 token 重試"
                    ])
                full_univ = None
    finally:
        os.environ["FINMIND_TOKEN"] = _orig_token_env

    if full_univ is None or len(full_univ) == 0:
        _safe_print("[sequential] ⚠️ 所有 token 都無法取得股票清單，中止")
        return

    total = len(full_univ)

    remaining_ids = set(full_univ["stock_id"].astype(str)) - already_scanned

    # Segment mode: explicit start/size takes priority (weighted multi-quota splits),
    # falling back to equal --batch-count division.
    _seg_idx = getattr(args, "batch_index", -1)
    _seg_cnt = max(1, getattr(args, "batch_count", 1))
    _seg_start = getattr(args, "segment_start", -1)
    _seg_size = getattr(args, "segment_size", -1)
    if _seg_start >= 0 and _seg_size > 0:
        _sorted_ids = sorted(full_univ["stock_id"].astype(str))
        _segment_ids = set(_sorted_ids[_seg_start: _seg_start + _seg_size])
        remaining_ids = remaining_ids & _segment_ids
        total = len(_segment_ids)
        _safe_print(
            f"[sequential] 加權分段 seg{_seg_idx}：第 {_seg_start}～{_seg_start + _seg_size - 1} 位，"
            f"共 {len(_segment_ids)} 支，剩餘 {len(remaining_ids)} 未掃"
        )
    elif _seg_idx >= 0 and _seg_cnt > 1:
        _sorted_ids = sorted(full_univ["stock_id"].astype(str))
        _sz = (len(_sorted_ids) + _seg_cnt - 1) // _seg_cnt
        _segment_ids = set(_sorted_ids[_seg_idx * _sz: (_seg_idx + 1) * _sz])
        remaining_ids = remaining_ids & _segment_ids
        total = len(_segment_ids)  # show segment total, not full 1800
        _safe_print(f"[sequential] 分段 {_seg_idx}/{_seg_cnt}：共 {len(_segment_ids)} 支，剩餘 {len(remaining_ids)} 未掃")
    _seg_label = f" [段{_seg_idx}]" if _seg_idx >= 0 else ""

    # Filter out stocks confirmed empty on a previous day (delisted / no price data ever)
    if no_data_ids:
        _before_nd = len(remaining_ids)
        remaining_ids -= no_data_ids
        _skipped_nd = _before_nd - len(remaining_ids)
        if _skipped_nd > 0:
            _safe_print(f"[sequential] 過濾 {_skipped_nd} 支歷史空資料股票（黑名單共 {len(no_data_ids)} 支）")

    # Filter out disposition + suspended stocks (risk management)
    _skip_ids: set[str] = set()
    try:
        _disp_ids = fetch_disposition_stocks(client, today)
        _skip_ids.update(_disp_ids)
    except Exception as _de:
        _safe_print(f"[sequential] ⚠️ 處置股資料取得失敗（無妨）：{_de}")
    try:
        _susp_ids = fetch_suspended_stocks(client, today)
        _skip_ids.update(_susp_ids)
    except Exception as _se:
        _safe_print(f"[sequential] ⚠️ 暫停交易資料取得失敗（無妨）：{_se}")
    if _skip_ids:
        _before_skip = len(remaining_ids)
        remaining_ids -= _skip_ids
        _skipped = _before_skip - len(remaining_ids)
        if _skipped > 0:
            _safe_print(f"[sequential] 過濾 {_skipped} 支處置/暫停交易股票（共 {len(_skip_ids)} 支在名單中）")

    _safe_print(f"[sequential] 今日已嘗試 {len(already_scanned)}/{total}，剩餘未掃 {len(remaining_ids)} 支")
    if os.getenv("DISCORD_WEBHOOK_URL"):
        _nd_note = f"（已過濾 `{len(no_data_ids)}` 支歷史無資料）" if no_data_ids else ""
        send_discord_messages([
            f"🚀 **掃描啟動{_seg_label}** · {_cst_now()} CST · {today}\n"
            f"今日已掃 `{len(already_scanned)}/{total}` | 本輪待掃 `{len(remaining_ids)}` 支 {_nd_note}"
        ])

    if not remaining_ids:
        msg = f"✅ **今日全部已掃完** · {_cst_now()} CST · {today}\n全部 `{total}` 支"
        _safe_print(msg)
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([msg])
        return

    token_list = [
        (0, os.getenv("FINMIND_TOKEN")),
        (1, os.getenv("FINMIND_TOKEN_2")),
        (2, os.getenv("FINMIND_TOKEN_3")),
        (3, os.getenv("FINMIND_TOKEN_4")),
        (4, os.getenv("FINMIND_TOKEN_5")),
        (5, os.getenv("FINMIND_TOKEN_6")),
        (6, os.getenv("FINMIND_TOKEN_7")),
        (7, os.getenv("FINMIND_TOKEN_8")),
        (8, os.getenv("FINMIND_TOKEN_9")),
    ]
    token_list = [(i, t) for i, t in token_list if t]
    _chunk_size = max(1, int(os.getenv("SEQUENTIAL_CHUNK_SIZE", "600") or "600"))
    # Detect duplicated tokens (same account key copied into multiple env vars).
    _token_to_idxs: dict[str, list[int]] = {}
    for _i, _t in token_list:
        _token_to_idxs.setdefault(str(_t), []).append(_i)
    _dup_groups = [idxs for idxs in _token_to_idxs.values() if len(idxs) > 1]
    if _dup_groups:
        _dup_msg = "；".join([",".join([str(x) for x in g]) for g in _dup_groups])
        _safe_print(f"[sequential] ⚠️ 偵測到重複 token（帳號索引：{_dup_msg}），可能不是三個不同帳號")
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([
                f"⚠️ **偵測到重複 FINMIND_TOKEN** · {_cst_now()} CST\n"
                f"帳號索引 `{_dup_msg}` 使用相同 token，請確認三個帳號是否不同"
            ])

    round_new = 0
    zero_quota_accounts: set[int] = set()
    orig_env_token = os.environ.get("FINMIND_TOKEN", "")

    _env_key_map = {
        0: "FINMIND_TOKEN",   1: "FINMIND_TOKEN_2", 2: "FINMIND_TOKEN_3",
        3: "FINMIND_TOKEN_4", 4: "FINMIND_TOKEN_5", 5: "FINMIND_TOKEN_6",
        6: "FINMIND_TOKEN_7", 7: "FINMIND_TOKEN_8", 8: "FINMIND_TOKEN_9",
    }

    # ── 掃描前一次預檢所有帳號配額，結果快取供迴圈重用（避免重複消耗 API 呼叫）──
    _quota_probe_cache: dict[int, tuple[bool, str]] = {}
    _quota_lines: list[str] = []
    for _pi, _pt in token_list:
        _pk = _env_key_map.get(_pi, f"FINMIND_TOKEN_{_pi}")
        if _pk in _quota_exhausted_keys:
            _quota_probe_cache[_pi] = (False, "取清單時疑似配額不足（僅提示）")
            _quota_lines.append(f"帳號 {_pi}: ⚠️ 疑似不足（取清單）")
            continue
        os.environ["FINMIND_TOKEN"] = _pt
        _pc = FinMindClient(cache_dir=client.cache_dir)
        _pok, _pmsg = probe_batch_quota(_pc)
        os.environ["FINMIND_TOKEN"] = orig_env_token
        _quota_probe_cache[_pi] = (_pok, _pmsg)
        _quota_lines.append(f"帳號 {_pi}: {'✅ 可用（探測）' if _pok else '⚠️ 疑似不足（探測）'}")
    _safe_print(f"[sequential] 帳號配額預檢：{' | '.join(_quota_lines)}")
    if os.getenv("DISCORD_WEBHOOK_URL"):
        send_discord_messages([
            f"📊 **帳號配額預檢** · {_cst_now()} CST\n" + "\n".join(_quota_lines)
        ])

    # Multi-pass budget:
    # If precheck says N accounts are available, run up to N passes in the same job
    # (i.e., re-run whole sequential flow N-1 extra times) to improve handoff success.
    _pre_ok_cnt = sum(1 for _ok, _ in _quota_probe_cache.values() if _ok)
    _planned_budget = 0  # single pass only — no recursive re-run within the same job
    _pass_budget = int(getattr(args, "_sequential_reentry_budget", _planned_budget) or 0)

    for token_idx, token in token_list:
        if not remaining_ids:
            break
        if token_idx in zero_quota_accounts:
            _safe_print(f"[sequential] 帳號 {token_idx}: 本輪已確認配額耗盡，略過")
            continue
        _remaining_before = len(remaining_ids)

        # Never hard-skip a token in sequential mode based on pre-probe signals.
        # Each account always gets a real scan attempt.

        # 直接切換到下一個帳號，不做固定等待。
        # 若遇到 IP 限流，交由下方 probe + retry 機制處理。
        os.environ["FINMIND_TOKEN"] = token
        token_client = FinMindClient(cache_dir=client.cache_dir)

        # 每次帳號正式開掃前都重新探測一次該帳號配額（使用最新狀態）。
        # 注意：此探測結果只做提示，不作硬性跳過。實務上 FinMind 在帳號切換
        # 當下可能回傳暫時性 402，若直接略過會錯失可用帳號。
        _pre_ok, _pre_msg = probe_batch_quota(token_client)
        if not _pre_ok:
            _is_quota_issue = "配額已耗盡" in _pre_msg or "upper limit" in _pre_msg.lower()
            _is_ip_throttle = "IP 限流" in _pre_msg or "ip" in _pre_msg.lower() and "限流" in _pre_msg
            if _is_quota_issue:
                _safe_print(f"[sequential] 帳號 {token_idx}: 開掃前配額檢查顯示不足（{_pre_msg}），仍嘗試實掃確認")
                if os.getenv("DISCORD_WEBHOOK_URL"):
                    send_discord_messages([
                        f"⚠️ **帳號 {token_idx} 開掃前配額疑似不足** · {_cst_now()} CST\n"
                        f"{_pre_msg}\n改以實際掃描再確認，避免誤判略過"
                    ])
            elif _is_ip_throttle:
                _safe_print(f"[sequential] 帳號 {token_idx}: 開掃前偵測到 IP 限流（{_pre_msg}），稍後仍嘗試實掃")
                if os.getenv("DISCORD_WEBHOOK_URL"):
                    send_discord_messages([
                        f"⚠️ **帳號 {token_idx} 開掃前 IP 限流** · {_cst_now()} CST\n"
                        f"{_pre_msg}\n將嘗試繼續掃描"
                    ])



        # 不再用切換前 probe 決定跳過，避免誤判「有額度帳號」為無額度。
        # 直接進入實際掃描；若途中真的配額耗盡，build_daily_snapshot 會回報
        # quota_exhausted，並由後續邏輯切換下一帳號。
        n_rem = len(remaining_ids)
        _safe_print(f"[sequential] 帳號 {token_idx}: 額度正常，開始掃描 {n_rem} 支")
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([
                f"🔄 **帳號 {token_idx} 開始{_seg_label}** · {_cst_now()} CST\n"
                f"剩餘 `{n_rem}/{total}` 支待掃"
            ])

        remaining_univ_all = full_univ[full_univ["stock_id"].astype(str).isin(remaining_ids)]
        remaining_univ = remaining_univ_all.head(_chunk_size)
        if len(remaining_univ_all) > len(remaining_univ):
            _safe_print(
                f"[sequential] 帳號 {token_idx}: 本輪採用 chunk 掃描 {len(remaining_univ)} 支"
                f"（尚有 {len(remaining_univ_all) - len(remaining_univ)} 支待後續接力）"
            )
        gap_file = scan_dir / f"_seq_{token_idx}.csv"
        remaining_univ.to_csv(gap_file, index=False)

        token_csv = scan_dir / f"batch_seq{token_idx}_{today}.csv"
        _ckpt_csv = scan_dir / f"_ckpt_seq{token_idx}_{today}.csv"
        _ckpt_csv.unlink(missing_ok=True)  # start fresh for this scan run

        scan_args = _copy.copy(args)
        scan_args.stocks = str(gap_file)
        scan_args.batch_index = -1
        scan_args.notify = False
        scan_args.checkpoint_csv = str(_ckpt_csv)

        _scan_failed = False
        _scan_exc = None
        try:
            candidates, watchlist, _, breadth = build_daily_snapshot(scan_args, token_client, config)
        except Exception as _e:
            _scan_failed = True
            _scan_exc = _e
        finally:
            gap_file.unlink(missing_ok=True)
            os.environ["FINMIND_TOKEN"] = orig_env_token

        if _scan_failed:
            _exc_str = str(_scan_exc).lower() if _scan_exc is not None else ""
            _is_quota = (
                "402" in _exc_str
                or "upper limit" in _exc_str
                or "配額" in _exc_str
                or "payment required" in _exc_str
            )
            if _is_quota:
                _safe_print(f"[sequential] 帳號 {token_idx}: 掃描中配額耗盡（{_scan_exc}），切換下一帳號")
                if os.getenv("DISCORD_WEBHOOK_URL"):
                    send_discord_messages([
                        f"⏰ **帳號 {token_idx} 掃描中配額耗盡** · {_cst_now()} CST\n"
                        f"切換下一帳號繼續"
                    ])
            else:
                _safe_print(f"[sequential] 帳號 {token_idx}: 掃描異常（{_scan_exc}），切換下一帳號")
            continue

        _quota_hit = bool(breadth.get("quota_exhausted"))
        n_actual = int(breadth.get("total_stocks", 0))
        if _quota_hit and n_actual == 0 and len(remaining_univ) > 0:
            _safe_print(f"[sequential] 帳號 {token_idx}: 首次掃描即 0 支且回報配額耗盡，立即重試一次")
            os.environ["FINMIND_TOKEN"] = token
            token_client = FinMindClient(cache_dir=client.cache_dir)
            gap_file = scan_dir / f"_seq_{token_idx}.csv"
            remaining_univ.to_csv(gap_file, index=False)
            _ckpt_csv = scan_dir / f"_ckpt_seq{token_idx}_{today}.csv"
            _ckpt_csv.unlink(missing_ok=True)
            scan_args.checkpoint_csv = str(_ckpt_csv)
            try:
                candidates, watchlist, _, breadth = build_daily_snapshot(scan_args, token_client, config)
            except Exception as _scan_exc2:
                _safe_print(f"[sequential] 帳號 {token_idx}: 重試仍失敗（{_scan_exc2}），切換下一帳號")
                gap_file.unlink(missing_ok=True)
                os.environ["FINMIND_TOKEN"] = orig_env_token
                continue
            finally:
                gap_file.unlink(missing_ok=True)
                os.environ["FINMIND_TOKEN"] = orig_env_token

        # Determine which stocks were actually attempted this scan.
        # breadth["scanned_ids"] only contains stocks that returned price data
        # (i.e., entered signals_by_stock). Stocks with empty prices, short history,
        # or FinMind returning no rows are NOT in scanned_ids — but they WERE
        # attempted and should not be re-scanned next hour.
        #
        # Fix: if quota was NOT exhausted, every stock in remaining_univ was
        # attempted (some just had no data). Mark the full set as done.
        # If quota WAS exhausted, only the fetched stocks are safe to mark as done;
        # the rest were skipped by the quota event and must be retried.
        _quota_hit = bool(breadth.get("quota_exhausted"))
        n_actual = int(breadth.get("total_stocks", 0))
        scanned_now: set[str] = set(breadth.get("scanned_ids") or set())
        _success_ratio = n_actual / max(len(remaining_univ), 1)
        _ip_throttled = (not _quota_hit and n_actual > 0 and _success_ratio < 0.5)
        if not _quota_hit and n_actual > 0:
            if _ip_throttled:
                # Success rate < 50% → likely IP throttling mid-scan.
                # Only mark stocks that actually returned data as done;
                # leave un-fetched stocks for the next account to retry.
                _throttled_count = len(remaining_univ) - len(scanned_now)
                _safe_print(
                    f"[sequential] 帳號 {token_idx}: 成功率低 ({_success_ratio:.0%}，"
                    f"{n_actual}/{len(remaining_univ)})，疑似 IP 限流，"
                    f"{_throttled_count} 支未掃留給下一帳號重試"
                )
                if os.getenv("DISCORD_WEBHOOK_URL"):
                    send_discord_messages([
                        f"⚠️ **帳號 {token_idx} IP 限流** · {_cst_now()} CST\n"
                        f"成功率 {_success_ratio:.0%} ({n_actual}/{len(remaining_univ)})，"
                        f"{_throttled_count} 支未掃，交由下一帳號重試"
                    ])
                # scanned_now already contains only scanned_ids — no further update needed
            else:
                # Normal completion: mark all remaining_univ as attempted.
                scanned_now.update(remaining_univ["stock_id"].astype(str))

                # Stocks that returned NO price data → persistent blacklist.
                _got_data = set(breadth.get("scanned_ids") or set())
                _new_no_data = set(remaining_univ["stock_id"].astype(str)) - _got_data
                if _new_no_data:
                    no_data_ids.update(_new_no_data)
                    pd.DataFrame({"stock_id": sorted(no_data_ids)}).to_csv(
                        no_data_csv, index=False, encoding="utf-8-sig"
                    )
                    _safe_print(
                        f"[sequential] 帳號 {token_idx}: 新增 {len(_new_no_data)} 支到空資料黑名單"
                        f"（黑名單共 {len(no_data_ids)} 支）"
                    )
        elif not _quota_hit and n_actual == 0 and len(remaining_univ) > 0:
            _safe_print(
                f"[sequential] 帳號 {token_idx}: 本輪 0 支有效掃描，"
                "不將剩餘股票標記為已嘗試，交給下一帳號/下一輪"
            )

        if scanned_now:
            already_scanned.update(scanned_now)
            remaining_ids -= scanned_now
            round_new += len(scanned_now)

            # Persist attempted IDs immediately so next hourly run skips these stocks
            pd.DataFrame({"stock_id": sorted(already_scanned)}).to_csv(
                attempted_csv, index=False, encoding="utf-8-sig"
            )

        # Save this token's raw results immediately.
        # Primary: incremental checkpoint written stock-by-stock during scan (preserves data
        # even if the process crashes or quota exhausts mid-scan).
        # Fallback: in-memory full snapshot, then candidates/watchlist.
        if _ckpt_csv.exists():
            raw_df = pd.read_csv(_ckpt_csv, encoding="utf-8-sig").drop_duplicates(subset=["stock_id"])
            _ckpt_csv.unlink(missing_ok=True)
        else:
            full_snapshot = breadth.pop("_snapshot", pd.DataFrame())
            if not full_snapshot.empty:
                raw_df = full_snapshot.copy()
            elif not watchlist.empty:
                raw_df = pd.concat([candidates, watchlist], ignore_index=True)
            else:
                raw_df = candidates.copy()
        if token_csv.exists():
            try:
                if token_csv.stat().st_size > 0:
                    prev = pd.read_csv(token_csv, encoding="utf-8-sig")
                    if not prev.empty:
                        raw_df = pd.concat([prev, raw_df], ignore_index=True).drop_duplicates(subset=["stock_id"])
            except Exception as _csv_exc:
                _safe_print(f"[sequential] ⚠️ 無法讀取 {token_csv.name}（{_csv_exc}），略過合併")
        if not raw_df.empty:
            raw_df.to_csv(token_csv, index=False, encoding="utf-8-sig")
            _write_summary_excel(raw_df, token_csv.with_suffix(".xlsx"))

        # ── Notion 即時同步（掃描段落完成後立即上傳有訊號股票）────────────────────────
        if notion_enabled() and not raw_df.empty and (
            not candidates.empty or not watchlist.empty
        ):
            try:
                _notion_date = str(raw_df["date"].max())[:10] if "date" in raw_df.columns else today
                _regime_str = str(breadth.get("market_regime", "") if isinstance(breadth, dict) else "")
                sync_scan_results(
                    candidates if not candidates.empty else pd.DataFrame(),
                    watchlist if not watchlist.empty else pd.DataFrame(),
                    _notion_date,
                    {},
                    market_regime=_regime_str,
                )
                _safe_print(
                    f"[sequential] ✅ Notion 即時同步完成"
                    f"（帳號 {token_idx}，候選 {len(candidates)}，觀察 {len(watchlist)}）"
                )
            except Exception as _notion_exc:
                _safe_print(f"[sequential] ⚠️ Notion 同步失敗（skip）：{_notion_exc}")

        _remaining_after = len(remaining_ids)
        _delta_done = _remaining_before - _remaining_after

        if raw_df.empty and len(remaining_univ) > 0:
            _safe_print(f"[sequential] 帳號 {token_idx}: ⚠️ 掃描 0 支有效資料（IP 仍在限流中），本輪略過")
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([
                    f"⚠️ **帳號 {token_idx} 掃描 0 支資料** · {_cst_now()} CST\n"
                    f"IP 仍在限流中或配額不足，已略過儲存，下輪繼續接力"
                ])

        _cand_n = len(candidates)
        _watch_n = len(watchlist)
        if _quota_hit:
            if n_actual == 0:
                zero_quota_accounts.add(token_idx)
            _safe_print(f"[sequential] 帳號 {token_idx}: ⚡ 配額耗盡，掃到 {n_actual} 支後提前中止，切換下一帳號")
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([
                    f"⚡ **帳號 {token_idx} 配額耗盡** · {_cst_now()} CST\n"
                    f"掃到 `{n_actual}` 支後中止，切換下一帳號\n"
                    f"（若三帳號皆立即耗盡，請確認三個 FINMIND_TOKEN 是否為不同帳號）"
                ])
        elif _ip_throttled:
            _safe_print(f"[sequential] 帳號 {token_idx}: IP 限流，直接切換下一帳號")
        elif n_actual == 0 and len(remaining_univ) > 0:
            _safe_print(f"[sequential] 帳號 {token_idx}: ⚠️ 實際掃 0 支（API 可能靜默限流），本輪略過儲存")
        _safe_print(
            f"[sequential] 帳號 {token_idx}: 完成，實際掃 {n_actual} 支，"
            f"候選 {_cand_n}，觀察 {_watch_n}，"
            f"本帳號接續完成 {_delta_done} 支，累計嘗試 {len(already_scanned)}/{total}"
        )
        if os.getenv("DISCORD_WEBHOOK_URL"):
            _status = "✅" if _cand_n > 0 else "🔵"
            _quota_note = " ⚡配額耗盡" if _quota_hit else ""
            send_discord_messages([
                f"{_status} **帳號 {token_idx} 完成{_quota_note}** · {_cst_now()} CST\n"
                f"掃 `{n_actual}` 支 | 候選 `{_cand_n}` | 觀察 `{_watch_n}`\n"
                f"接續完成 `{_delta_done}` 支（{_remaining_before} → {_remaining_after}）\n"
                f"本段累計 `{len(already_scanned)}/{total}` 支"
            ])

    # If all accounts returned 0 results, decide whether to mark as done based on
    # day-of-week. On weekends/holidays FinMind has no data for the whole day →
    # mark remaining as attempted so relay doesn't retry indefinitely.
    # On weekdays, 0 results usually means the scan ran too early (before FinMind
    # updates EOD data, e.g. midnight) → do NOT mark, just exit so the next
    # scheduled run (07:30+ CST) can retry with data available.
    if round_new == 0 and remaining_ids:
        import datetime as _dt
        _cst_now_dt = _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=8)))
        _is_weekend = _cst_now_dt.weekday() >= 5  # Saturday=5, Sunday=6
        if _is_weekend:
            _safe_print(
                f"[sequential] ⚠️ 週末/假日：本輪所有帳號均返回 0 支有效資料，"
                f"將剩餘 {len(remaining_ids)} 支標記為今日已嘗試（明日重新掃）"
            )
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([
                    f"📅 **週末/假日無資料** · {_cst_now()} CST\n"
                    f"剩餘 `{len(remaining_ids)}` 支標記為今日已嘗試，明日重新開始"
                ])
            already_scanned.update(remaining_ids)
            remaining_ids.clear()
            pd.DataFrame({"stock_id": sorted(already_scanned)}).to_csv(
                attempted_csv, index=False, encoding="utf-8-sig"
            )
        else:
            # Weekday: 0 results = scan ran too early or transient API issue.
            # Do NOT pre-mark stocks as done — next scheduled run will retry.
            _safe_print(
                f"[sequential] ⚠️ 平日本輪 0 支有效資料（時間太早或 API 暫時無資料），"
                f"不標記已嘗試，下次接力繼續"
            )
            if os.getenv("DISCORD_WEBHOOK_URL"):
                send_discord_messages([
                    f"⚠️ **本輪 0 支資料（平日）** · {_cst_now()} CST\n"
                    f"可能是時間太早（FinMind 尚未更新）或 API 暫時異常\n"
                    f"不標記已嘗試，等待下次接力自動重試"
                ])
            remaining_ids.clear()  # 結束本輪迴圈，但不持久化

    # If this round made progress and reentry budget remains, run another full pass.
    # Example:
    #   precheck 2 accounts available -> budget 1 (run once more)
    #   precheck 3 accounts available -> budget 2 (run twice more)
    if remaining_ids and round_new > 0 and _pass_budget > 0:
        _safe_print(
            f"[sequential] 本輪有進度 {round_new} 支，剩餘 {len(remaining_ids)} 支，"
            f"啟動整輪重跑（剩餘重跑次數 {_pass_budget}）"
        )
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([
                f"🔁 **啟動一次整輪重跑** · {_cst_now()} CST\n"
                f"本輪新掃 `{round_new}` 支，仍剩 `{len(remaining_ids)}` 支\n"
                f"依預檢可用帳號策略，將再重跑一次（剩餘重跑次數 `{_pass_budget}`）"
            ])
        rerun_args = _copy.copy(args)
        setattr(rerun_args, "_sequential_reentry_budget", _pass_budget - 1)
        run_sequential_scan(rerun_args, client, config)
        return

    # Round summary
    _pct = f"{len(already_scanned) / max(total, 1):.0%}"
    if not remaining_ids:
        msg = (
            f"🎉 **今日全部掃完{_seg_label}！** · {_cst_now()} CST · {today}\n"
            f"全部 `{total}` 支已覆蓋（100%） | 本輪新掃 `{round_new}` 支"
        )
    else:
        _passes_left = max(1, round(len(remaining_ids) / max(round_new, 1)))
        msg = (
            f"⏸ **本輪結束{_seg_label}** · {_cst_now()} CST · {today}\n"
            f"本輪新掃 `{round_new}` 支 | 累計 `{len(already_scanned)}/{total}`（{_pct}）\n"
            f"剩餘 `{len(remaining_ids)}` 支 · 預估還需約 `{_passes_left}` 輪（每2小時自動接力）"
        )
    _safe_print(msg)
    if os.getenv("DISCORD_WEBHOOK_URL"):
        send_discord_messages([msg])

    # Rename batch_seq files whose name uses the run date but data is from the prior trading day.
    for _bsf in sorted(scan_dir.glob(f"batch_seq*_{today}.csv")):
        try:
            if _bsf.stat().st_size == 0:
                continue
            _bsdf = pd.read_csv(_bsf, encoding="utf-8-sig")
            if "date" not in _bsdf.columns or _bsdf.empty:
                continue
            _real = str(pd.to_datetime(_bsdf["date"]).max().strftime("%Y-%m-%d"))
            if _real >= today or _real <= "2020-01-01":
                continue
            _dst = _bsf.parent / _bsf.name.replace(today, _real)
            if _dst.exists() and _dst.stat().st_size > 0:
                _old = pd.read_csv(_dst, encoding="utf-8-sig")
                _bsdf = pd.concat([_bsdf, _old], ignore_index=True).drop_duplicates(subset=["stock_id"], keep="first")
            _bsdf.to_csv(_dst, index=False, encoding="utf-8-sig")
            _write_summary_excel(_bsdf, _dst.with_suffix(".xlsx"))
            _bsf.unlink(missing_ok=True)
            _bsf.with_suffix(".xlsx").unlink(missing_ok=True)
            _safe_print(f"[sequential] 📅 {_bsf.name} → {_dst.name} ({len(_bsdf)} 支)")
        except Exception as _e:
            _safe_print(f"[sequential] ⚠️ 重命名失敗（無妨）：{_bsf.name} → {_e}")


def run_smart_scan(args: argparse.Namespace, config: StrategyConfig) -> None:
    """智能掃描：先顯示進度，依序檢查三個帳號配額，有額度就掃，沒有就換下一個。"""
    batch_dir = Path(args.output) / "full_scan"
    batch_dir.mkdir(parents=True, exist_ok=True)
    cache_dir = Path(args.output) / "cache"

    # 1. 讀取已掃進度
    scanned_ids: set[str] = set()
    for p in sorted(batch_dir.glob("batch_*.csv")):
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if "stock_id" in df.columns:
                scanned_ids.update(df["stock_id"].astype(str))
        except Exception:
            pass

    total_est = args.max_universe
    pct = f"{len(scanned_ids) / total_est:.0%}" if total_est else "?"
    progress_line = f"已掃 `{len(scanned_ids)}` 支（約 `{pct}` of {total_est}）"

    # 2. 依序檢查三個帳號配額
    tokens_config = [
        ("FINMIND_TOKEN",   "L"),
        ("FINMIND_TOKEN_2", "950223"),
        ("FINMIND_TOKEN_3", "rekis"),
    ]
    quota_results: list[tuple[str, str, bool, str]] = []  # (env_key, name, ok, token)
    status_lines: list[str] = []
    for env_key, name in tokens_config:
        token = os.getenv(env_key, "").strip()
        if not token:
            status_lines.append(f"• {name}：❌ 未設定")
            quota_results.append((env_key, name, False, ""))
            continue
        os.environ["FINMIND_TOKEN"] = token
        tmp_client = FinMindClient(cache_dir=cache_dir)
        quota_ok, _ = probe_batch_quota(tmp_client)
        icon = "✅ 有額度" if quota_ok else "⏰ 配額耗盡"
        status_lines.append(f"• {name}：{icon}")
        quota_results.append((env_key, name, quota_ok, token))

    summary = "\n".join([
        f"📋 **掃描狀態** · {_cst_now()} CST · {args.end}",
        progress_line,
        *status_lines,
    ])
    _safe_print(summary)
    if os.getenv("DISCORD_WEBHOOK_URL"):
        send_discord_messages([summary])

    # 3. 依序用有額度的帳號掃描
    original_token = os.environ.get("FINMIND_TOKEN", "")
    try:
        for env_key, account_name, quota_ok, token in quota_results:
            if not quota_ok or not token:
                _safe_print(f"[smart-scan] 跳過 {account_name}（無額度或未設定）")
                continue
            _safe_print(f"[smart-scan] 開始使用 {account_name} 帳號")
            os.environ["FINMIND_TOKEN"] = token
            account_client = FinMindClient(cache_dir=cache_dir)
            run_continue_scan(args, account_client, config)
    finally:
        os.environ["FINMIND_TOKEN"] = original_token


def run_continue_scan(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    """掃描尚未掃描的股票，直到 API 配額耗盡為止。

    從 output/full_scan/batch_*.csv 讀取已掃記錄，對剩餘股票繼續掃描，
    儲存到下一個可用的 batch_NN.csv/xlsx。
    """
    import copy

    token_ok, token_msg = validate_finmind_token()
    if not token_ok:
        err = f"❌ **繼續掃描中止：Token 無效** · {_cst_now()} CST\n{token_msg}"
        _safe_print(err)
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([err])
        return

    quota_ok, quota_msg = probe_batch_quota(client)
    if not quota_ok:
        _safe_print(f"[continue-scan] 配額不足，跳過此輪。{quota_msg}")
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([f"⏰ **繼續掃描跳過** · {_cst_now()} CST\n{quota_msg}"])
        return

    batch_dir = Path(args.output) / "full_scan"
    batch_dir.mkdir(parents=True, exist_ok=True)

    scanned_ids: set[str] = set()
    for p in sorted(batch_dir.glob("batch_*.csv")):
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if "stock_id" in df.columns:
                scanned_ids.update(df["stock_id"].astype(str))
        except Exception:
            pass

    stock_info = fetch_stock_info(client)
    full_universe = build_auto_universe(stock_info, max_symbols=args.max_universe)
    all_ids = set(full_universe["stock_id"].astype(str))
    remaining_ids = all_ids - scanned_ids
    total_universe = len(all_ids)
    total_scanned = len(scanned_ids & all_ids)

    if not remaining_ids:
        _safe_print(f"[continue-scan] 全部掃完！{total_scanned}/{total_universe} 支")
        if os.getenv("DISCORD_WEBHOOK_URL"):
            send_discord_messages([
                f"✅ **全市場掃描完成** · {_cst_now()} CST · {args.end}\n"
                f"已掃 {total_scanned}/{total_universe} 支"
            ])
        return

    remaining = full_universe[full_universe["stock_id"].astype(str).isin(remaining_ids)].reset_index(drop=True)
    pct = f"{total_scanned / total_universe:.0%}" if total_universe else "0%"
    _safe_print(f"[continue-scan] 尚未掃描 {len(remaining)} 支（已掃 {total_scanned}/{total_universe}，{pct}）")
    if os.getenv("DISCORD_WEBHOOK_URL"):
        send_discord_messages([
            f"🔄 **繼續掃描** · {_cst_now()} CST · {args.end}\n"
            f"尚餘 `{len(remaining)}` 支 · 已掃 `{total_scanned}/{total_universe}`（`{pct}`）"
        ])

    existing = sorted(batch_dir.glob("batch_[0-9][0-9].csv"))
    next_num = int(existing[-1].stem.split("_")[-1]) + 1 if existing else 0

    _tmp = batch_dir / "_continue_tmp.csv"
    remaining[["stock_id", "name"]].to_csv(_tmp, index=False, encoding="utf-8-sig")
    scan_args = copy.copy(args)
    scan_args.stocks = str(_tmp)
    scan_args.batch_index = -1

    try:
        candidates, watchlist, _, breadth = build_daily_snapshot(scan_args, client, config)
    finally:
        _tmp.unlink(missing_ok=True)

    newly_scanned = int(breadth.get("total_stocks", 0))
    batch_csv = batch_dir / f"batch_{next_num:02d}.csv"
    batch_xlsx = batch_dir / f"batch_{next_num:02d}.xlsx"
    batch_df = pd.concat([candidates, watchlist], ignore_index=True) if not watchlist.empty else candidates
    batch_df = batch_df.sort_values("entry_score", ascending=False).drop_duplicates(subset=["stock_id"])
    batch_df.to_csv(batch_csv, index=False, encoding="utf-8-sig")
    batch_df.to_excel(batch_xlsx, index=False, engine="openpyxl")
    _safe_print(f"[continue-scan] 本輪掃 {newly_scanned} 支 → {batch_csv.name}")

    if os.getenv("DISCORD_WEBHOOK_URL"):
        new_total = total_scanned + newly_scanned
        remaining_after = max(0, total_universe - new_total)
        _cand_n = len(candidates)
        _watch_n = len(watchlist)
        _status = "✅" if _cand_n > 0 else "🔵"
        send_discord_messages([
            f"{_status} **批次 {next_num:02d} 完成** · {_cst_now()} CST · {args.end}\n"
            f"候選 `{_cand_n}` 檔 | 觀察 `{_watch_n}` 檔 | 本輪掃 `{newly_scanned}` 檔\n"
            f"進度：`{new_total}/{total_universe}` 支（剩 `{remaining_after}` 支）"
        ])


def run_hybrid_monitor(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    if args.stocks != "auto":
        universe = load_universe(args, client)
        candidates = pd.DataFrame(columns=["stock_id", "name"])
        watchlist = universe.copy()
        watch_pool = universe.copy()
        breadth: dict[str, object] = {}
    else:
        candidates, watchlist, universe, breadth = build_daily_snapshot(args, client, config)
        watch_pool = candidates.copy()
        if len(watch_pool) < args.watch_top:
            extra = watchlist.head(args.watch_top - len(watch_pool))
            watch_pool = pd.concat([watch_pool, extra], ignore_index=True)
    watch_symbols = watch_pool["stock_id"].astype(str).head(args.watch_top).tolist()

    fugle = FugleClient()
    if fugle.enabled and watch_symbols:
        live_quotes = fetch_watch_quotes(fugle, watch_symbols)
    else:
        live_quotes = pd.DataFrame(
            [{"symbol": symbol, "error": "FUGLE_API_KEY not configured"} for symbol in watch_symbols]
        )

    report_path = save_hybrid_report(args.output, candidates, watchlist, live_quotes, universe)
    news_map = {}
    if args.include_news:
        news_map = build_news_map(args.output, watch_pool.to_dict("records"), news_limit=args.news_limit)
    actual_date = str(breadth.get("actual_date") or args.end) if breadth else args.end
    message = format_hybrid_message_rich(candidates, watchlist, live_quotes, actual_date, news_map=news_map)
    _safe_print(message)
    _safe_print("")
    _safe_print(f"Hybrid report: {report_path}")
    if args.notify:
        send_discord_messages(split_message(message))


def run_sponsor_monitor(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    if args.stocks != "auto":
        universe = load_universe(args, client)
        candidates = pd.DataFrame(columns=["stock_id", "name"])
        watchlist = universe.copy()
        watch_pool = universe.copy()
    else:
        candidates, watchlist, universe, _breadth = build_daily_snapshot(args, client, config)
        watch_pool = candidates.copy()
        if len(watch_pool) < args.watch_top:
            extra = watchlist.head(args.watch_top - len(watch_pool))
            watch_pool = pd.concat([watch_pool, extra], ignore_index=True)

    watch_symbols = watch_pool["stock_id"].astype(str).head(args.watch_top).tolist()

    last_report_path = None
    for cycle in range(1, args.repeat_count + 1):
        intraday_rows = collect_intraday_snapshot(client, watch_symbols, args.end)
        last_report_path = save_sponsor_monitor_report(args.output, candidates, watchlist, intraday_rows, universe)
        message = format_sponsor_message(candidates, watchlist, intraday_rows, args.end, cycle, args.repeat_count)
        _safe_print(message)
        _safe_print("")
        _safe_print(f"Sponsor report: {last_report_path}")
        if args.notify:
            send_discord_messages(split_message(message))
        if cycle < args.repeat_count:
            time.sleep(args.interval_seconds)


def run_event_monitor(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    token_ok, token_msg = validate_finmind_token()
    if not token_ok:
        err = f"❌ **爆量監控中止** · {args.end}\n{token_msg}"
        _safe_print(err)
        if args.notify:
            send_discord_messages([err])
        return

    if args.stocks != "auto":
        universe = load_universe(args, client)
        watch_pool = universe.copy()
        snapshot_lookup: dict[str, dict[str, object]] = {}
    else:
        try:
            candidates, watchlist, universe, _breadth = build_daily_snapshot(args, client, config)
        except Exception as exc:
            err = f"❌ **爆量監控中止** · {args.end}\n初始掃描失敗：{exc}"
            _safe_print(err)
            if args.notify:
                send_discord_messages([err])
            return
        watch_pool = candidates.copy()
        if len(watch_pool) < args.watch_top:
            extra = watchlist.head(args.watch_top - len(watch_pool))
            watch_pool = pd.concat([watch_pool, extra], ignore_index=True)
        watch_records = watch_pool.head(args.watch_top).to_dict("records")
        snapshot_lookup = {str(r["stock_id"]): r for r in watch_records}

    watch_pool = watch_pool.head(args.watch_top).copy()
    watch_symbols = watch_pool["stock_id"].astype(str).tolist()

    extra_symbols = [s.strip() for s in getattr(args, "watch_extra", "").split(",") if s.strip()]
    for sym in extra_symbols:
        if sym not in watch_symbols:
            watch_symbols.append(sym)

    fugle = FugleClient()
    if not fugle.enabled:
        msg = "⚠️ **爆量監控略過** · FUGLE_API_KEY 未設定，無法取得即時報價。\n請在 GitHub Secrets 加入 FUGLE_API_KEY 後重新啟動。"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    prev_cumvol: dict[str, float] = {}
    vol_history: dict[str, deque] = {}
    last_notified: dict[str, float] = {}
    consecutive_error_cycles = 0
    _MAX_ERROR_CYCLES = 10

    start_msg = (
        f"🟢 **爆量監控啟動** · {_cst_now()} CST\n"
        f"監控標的：{', '.join(watch_symbols)}\n"
        f"觸發條件：本分鐘量 ≥ 近{_VOL_WINDOW}分均量 × {args.event_volume_multiplier:.1f}倍"
    )
    _safe_print(start_msg)
    if args.notify:
        send_discord_messages(split_message(start_msg))

    for cycle in range(1, args.repeat_count + 1):
        quotes = fetch_watch_quotes(fugle, watch_symbols)

        all_errors = (
            not quotes.empty
            and "error" in quotes.columns
            and quotes["error"].notna().all()
        )
        if all_errors:
            sample_error = str(quotes["error"].iloc[0]) if not quotes.empty else ""
            is_auth_error = "401" in sample_error or "403" in sample_error or "金鑰" in sample_error
            is_rate_limit = "429" in sample_error or "頻率" in sample_error
            if is_rate_limit:
                _safe_print(f"[cycle {cycle}] 速率限制，等待 60 秒後繼續…")
                time.sleep(60)
            else:
                consecutive_error_cycles += 1
            if is_auth_error or consecutive_error_cycles >= _MAX_ERROR_CYCLES:
                reason = sample_error[:80] if sample_error else f"連續 {_MAX_ERROR_CYCLES} 個週期全部錯誤"
                abort_msg = (
                    f"⚠️ **爆量監控中止** · {_cst_now()} CST\n"
                    f"原因：{reason}\n"
                    f"請確認 FUGLE_API_KEY 是否有效，或手動重啟監控。"
                )
                _safe_print(abort_msg)
                if args.notify:
                    send_discord_messages(split_message(abort_msg))
                break
        else:
            consecutive_error_cycles = 0

        events, prev_cumvol = detect_volume_surges(
            quotes, prev_cumvol, vol_history, args.event_volume_multiplier
        )

        now_ts = time.time()
        for event in events:
            key = str(event["symbol"])
            if now_ts - last_notified.get(key, 0) < args.event_cooldown_seconds:
                continue
            message = format_volume_alert(event, snapshot_lookup)
            _safe_print("")
            _safe_print(message)
            if args.notify:
                send_discord_messages(split_message(message))
            last_notified[key] = now_ts

        if cycle < args.repeat_count:
            time.sleep(args.interval_seconds)


def run_daily_report(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    if args.notify:
        send_discord_messages([f"🔄 **盤後總結開始** · {_cst_now()} CST · {args.end}"])
    candidates, watchlist, universe, breadth = build_daily_snapshot(args, client, config)
    actual_date = str(breadth.get("actual_date") or args.end)
    watch_pool = candidates.head(args.watch_top).copy()
    if len(watch_pool) < args.watch_top:
        extra = watchlist.head(args.watch_top - len(watch_pool))
        watch_pool = pd.concat([watch_pool, extra], ignore_index=True)
    watch_symbols = watch_pool["stock_id"].astype(str).tolist()

    fugle = FugleClient()
    if fugle.enabled and watch_symbols:
        closing_quotes = fetch_watch_quotes(fugle, watch_symbols)
    else:
        closing_quotes = pd.DataFrame(
            [{"symbol": s, "error": "FUGLE_API_KEY not configured"} for s in watch_symbols]
        )

    news_map = build_news_map(args.output, watch_pool.to_dict("records"), news_limit=args.news_limit) if args.include_news else {}
    message = format_daily_report_message(candidates, watchlist, closing_quotes, actual_date, news_map=news_map, breadth=breadth)
    _safe_print(message)
    if args.notify:
        send_discord_messages(split_message(message))
    if notion_enabled():
        sync_scan_results(candidates, watchlist, actual_date, news_map, market_regime=str(breadth.get("market_regime") or ""))


def run_backtest_mode(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    stock_list = load_universe(args, client)
    market_raw = fetch_market_index(client, args.start, args.end)
    if market_raw.empty:
        raise RuntimeError("Unable to download TAIEX market data from FinMind.")
    market = prepare_market_frame(market_raw, config)
    signals_by_stock, signal_frames, _, _q = collect_signals(stock_list, client, market, config, args.start, args.end, args.workers)
    if not signals_by_stock:
        raise RuntimeError("No stock data was loaded. Check the stock universe and FinMind credentials.")

    backtest = run_backtest(signals_by_stock, market, config, args.capital)
    all_signals = pd.concat(signal_frames, ignore_index=True).sort_values(["date", "stock_id"])
    reports = save_reports(
        output_dir=args.output,
        metrics=backtest["metrics"],
        yearly=backtest["yearly"],
        trade_summary=backtest["trade_summary"],
        fills=backtest["fills"],
        equity_curve=backtest["equity_curve"],
        signals=all_signals,
        notes=backtest["notes"]
        + [
            "Daily backtest assumes signal evaluation and fills occur on the same close for simplicity.",
            "The earnings-date filter is optional because FinMind statement timing fields may vary by dataset.",
            "Position sizing uses integer shares and a 5% initial stop to cap risk at roughly 1% of portfolio equity.",
        ],
        config=config,
    )
    _safe_print("Backtest complete.")
    _safe_print(f"Excel report: {reports['excel']}")
    _safe_print(f"Equity chart: {reports['equity_chart']}")
    _safe_print(f"Yearly chart: {reports['yearly_chart']}")
    _safe_print("")
    _safe_print("Key metrics:")
    for key, value in backtest["metrics"].items():
        if isinstance(value, float):
            _safe_print(f"- {key}: {value:,.2f}")
        else:
            _safe_print(f"- {key}: {value}")


def run_walk_forward(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    """Run walk-forward validation: split the date range into folds and backtest each."""
    folds = getattr(args, "wf_folds", 4)
    overlap = getattr(args, "wf_overlap_days", 0)
    full_start = pd.Timestamp(args.start)
    full_end = pd.Timestamp(args.end)
    total_days = (full_end - full_start).days
    fold_days = total_days // folds

    _safe_print(f"Walk-forward: {folds} folds × ~{fold_days}d, {full_start.date()} → {full_end.date()}")

    market_raw = fetch_market_index(client, args.start, args.end)
    if market_raw.empty:
        raise RuntimeError("Unable to download TAIEX market data.")
    market_full = prepare_market_frame(market_raw, config)

    stock_list = load_universe(args, client)
    signals_by_stock, _, __, _q = collect_signals(stock_list, client, market_full, config, args.start, args.end, args.workers)
    if not signals_by_stock:
        raise RuntimeError("No signal data loaded.")

    fold_rows: list[dict[str, object]] = []
    _market_dates = pd.to_datetime(market_full["date"])
    _signal_dates = {sid: pd.to_datetime(frame["date"]) for sid, frame in signals_by_stock.items()}
    for fold_idx in range(folds):
        fold_start = full_start + pd.Timedelta(days=fold_idx * fold_days - overlap)
        fold_end = full_start + pd.Timedelta(days=(fold_idx + 1) * fold_days)
        if fold_idx == folds - 1:
            fold_end = full_end
        fold_start = max(fold_start, full_start)
        fold_label = f"{fold_start.date()} → {fold_end.date()}"

        market_slice = market_full[(_market_dates >= fold_start) & (_market_dates <= fold_end)].copy()
        signals_slice = {
            sid: frame[(_signal_dates[sid] >= fold_start) & (_signal_dates[sid] <= fold_end)].copy()
            for sid, frame in signals_by_stock.items()
        }
        signals_slice = {sid: f for sid, f in signals_slice.items() if not f.empty}
        if not signals_slice or market_slice.empty:
            _safe_print(f"  Fold {fold_idx + 1}: no data — skipping")
            continue

        try:
            result = run_backtest(signals_slice, market_slice, config, args.capital)
            m = result["metrics"]
            fold_rows.append({
                "fold": fold_idx + 1,
                "period": fold_label,
                "total_return_pct": m.get("total_return_pct", 0),
                "annual_return_pct": m.get("annual_return_pct", 0),
                "max_drawdown_pct": m.get("max_drawdown_pct", 0),
                "sharpe_ratio": m.get("sharpe_ratio", 0),
                "sortino_ratio": m.get("sortino_ratio", 0),
                "calmar_ratio": m.get("calmar_ratio", 0),
                "win_rate_pct": m.get("win_rate_pct", 0),
                "profit_factor": m.get("profit_factor", 0),
                "expectancy_pct": m.get("expectancy_pct", 0),
                "total_trades": m.get("total_trades", 0),
                "avg_holding_days": m.get("avg_holding_days", 0),
                "benchmark_return_pct": m.get("benchmark_return_pct", None),
                "alpha_pct": m.get("alpha_pct", None),
            })
            _safe_print(f"  Fold {fold_idx + 1} [{fold_label}]: "
                        f"return={m.get('total_return_pct', 0):+.1f}% "
                        f"DD={m.get('max_drawdown_pct', 0):.1f}% "
                        f"Sharpe={m.get('sharpe_ratio', 0):.2f} "
                        f"Sortino={m.get('sortino_ratio', 0):.2f} "
                        f"trades={m.get('total_trades', 0)}")
        except Exception as exc:
            _safe_print(f"  Fold {fold_idx + 1} error: {exc}")

    if fold_rows:
        wf_frame = pd.DataFrame(fold_rows)
        _safe_print("\nWalk-forward summary:")
        _safe_print(wf_frame.to_string(index=False))

        # Aggregate summary
        numeric_cols = ["total_return_pct", "max_drawdown_pct", "sharpe_ratio",
                        "sortino_ratio", "win_rate_pct", "profit_factor", "total_trades"]
        agg = {col: wf_frame[col].mean() for col in numeric_cols if col in wf_frame.columns}
        _safe_print("\nMean across folds:")
        for k, v in agg.items():
            _safe_print(f"  {k}: {v:.2f}")
        consistency = int((wf_frame["total_return_pct"] > 0).sum())
        _safe_print(f"  Positive folds: {consistency}/{len(fold_rows)}")

        # ── Enhanced metrics via backtest_pkg ────────────────────────────────────
        try:
            from backtest_pkg.metrics import confusion_matrix_summary, sharpe_ratio as _sharpe, max_drawdown as _mdd
            _returns = wf_frame["total_return_pct"].values / 100
            _eq_curve = np.cumprod(1 + _returns) * 100
            _sr = _sharpe(_returns, periods_per_year=len(_returns))
            _mdd_val = _mdd(_eq_curve)
            _win = int((wf_frame["total_return_pct"] > 0).sum())
            _n = len(wf_frame)
            _safe_print(f"\n[backtest_pkg] Cross-fold summary:")
            _safe_print(f"  Sharpe (annualised): {_sr:.3f}")
            _safe_print(f"  Max drawdown (fold equity): {_mdd_val:.1%}")
            _safe_print(f"  Win rate (folds): {_win}/{_n} = {_win/_n:.0%}")
        except Exception as _bm_exc:
            _safe_print(f"[backtest_pkg] metrics 計算失敗（skip）: {_bm_exc}")

        output_path = Path(args.output)
        output_path.mkdir(parents=True, exist_ok=True)
        wf_frame.to_csv(output_path / "walk_forward_results.csv", index=False)
        _safe_print(f"\nSaved: {output_path / 'walk_forward_results.csv'}")


def run_predict(args: argparse.Namespace, client: FinMindClient, config: StrategyConfig) -> None:
    token_ok, token_msg = validate_finmind_token()
    if not token_ok:
        msg = f"❌ **盤前預測中止** · {args.end}\n{token_msg}"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    today = args.end
    train_start = (pd.Timestamp(today) - pd.Timedelta(days=args.lookback_days)).strftime("%Y-%m-%d")

    try:
        market_df = fetch_market_index(client, train_start, today)
    except Exception as exc:
        msg = f"❌ **盤前預測失敗** · {today}\n無法取得大盤資料：{exc}"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    if market_df.empty:
        msg = f"⚠️ **盤前預測略過** · {today}\n大盤資料為空（FinMind 可能尚未更新）"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    # Fetch auxiliary data in parallel (all graceful-degrade on failure)
    with ThreadPoolExecutor(max_workers=8) as _pool:
        _f_us       = _pool.submit(fetch_us_features, train_start, today)
        _f_futures  = _pool.submit(fetch_futures_daily, client, train_start, today)
        _f_inst     = _pool.submit(fetch_futures_institutional, client, train_start, today)
        _f_calendar = _pool.submit(fetch_upcoming_events)
        _f_mkinst   = _pool.submit(fetch_market_institutional, client, train_start, today)
        _f_margin   = _pool.submit(fetch_market_margin, client, train_start, today)
        _f_pcr      = _pool.submit(fetch_options_pcr, client, train_start, today)
        _f_night    = _pool.submit(fetch_night_session, client, today)
        us_df           = _f_us.result()
        futures_raw     = _f_futures.result()
        inst_raw        = _f_inst.result()
        calendar_events = _f_calendar.result()
        mkinst_df       = _f_mkinst.result()
        margin_df       = _f_margin.result()
        pcr_df          = _f_pcr.result()
        night_data      = _f_night.result()

    futures_df = build_futures_features(market_df, futures_raw, inst_raw)

    # Merge 三大法人 + 融資融券 + PCR into a single inst_df for MarketPredictor
    inst_feat_df: pd.DataFrame | None = None
    _inst_frames = [df for df in (mkinst_df, margin_df, pcr_df) if not df.empty]
    if _inst_frames:
        inst_feat_df = _inst_frames[0]
        for _f in _inst_frames[1:]:
            inst_feat_df = inst_feat_df.merge(_f, on="date", how="outer")

    # Market-level news sentiment (Google RSS + FinMind combined)
    news_client = NewsClient(Path(args.output) / "news_cache")
    market_news = fetch_market_news_sentiment(news_client, finmind_client=client)

    predictor = MarketPredictor(horizon=5)
    predictor.fit(
        market_df,
        us_df        if not us_df.empty        else None,
        futures_df   if not futures_df.empty   else None,
        inst_feat_df if inst_feat_df is not None and not inst_feat_df.empty else None,
    )
    pred = predictor.predict_proba(
        market_df,
        us_df        if not us_df.empty        else None,
        futures_df   if not futures_df.empty   else None,
        inst_feat_df if inst_feat_df is not None and not inst_feat_df.empty else None,
    )

    # ── Log prediction to persistent history (PredictionStore) ───────────────────
    try:
        from storage.sqlite import PredictionStore
        from config.settings import get_settings
        _settings = get_settings()
        _pred_store = PredictionStore(_settings.db_path)
        _pred_store.log_prediction(
            date=today,
            prob_up=pred["prob_up"],
            predicted_up=pred["prob_up"] >= 0.5,
            model="market_xgb",
            stock_id="TAIEX",
        )
        _safe_print(f"[predict] 預測紀錄已存入 PredictionStore (prob_up={pred['prob_up']:.3f})")
    except Exception as _ps_exc:
        _safe_print(f"[predict] PredictionStore 記錄失敗（skip）: {_ps_exc}")

    # Build structured market_data JSON for Claude scenario analysis
    def _safe_float(val: object) -> object:
        import math as _math
        try:
            f = float(val)  # type: ignore[arg-type]
            return None if (_math.isnan(f) or _math.isinf(f)) else round(f, 4)
        except (TypeError, ValueError):
            return None

    _us_row = us_df.iloc[-1] if not us_df.empty else {}
    _ft_row = futures_df.iloc[-1] if not futures_df.empty else {}
    _mk_row = mkinst_df.iloc[-1] if not mkinst_df.empty else {}
    _pcr_val = float(pcr_df["pcr"].iloc[-1]) if not pcr_df.empty else None

    market_data_json: dict = {
        "us_market": {
            "nasdaq_ret": _safe_float(_us_row.get("nasdaq_ret1")),
            "sox_ret":    _safe_float(_us_row.get("sox_ret1")),
            "sp500_ret":  _safe_float(_us_row.get("sp500_ret1")),
            "tsm_adr_ret": _safe_float(_us_row.get("tsm_adr_ret1")),
            "nvda_ret":   _safe_float(_us_row.get("nvda_ret1")),
            "vix":        _safe_float(_us_row.get("vix")),
            "us10y_ret":  _safe_float(_us_row.get("us10y_ret1")),
            "dxy_ret":    _safe_float(_us_row.get("dxy_ret1")),
        },
        "night_session": night_data if night_data else None,
        "foreign_capital": {
            "futures_net":  _safe_float(_ft_row.get("foreign_futures_net")),
            "cash_norm":    _safe_float(_mk_row.get("foreign_inst_norm")),
            "pcr":          round(_pcr_val, 3) if _pcr_val is not None else None,
        },
        "xgb_prob_up": round(pred["prob_up"], 3),
        "xgb_label":   pred["label"],
    }
    scenario_text = generate_scenario_analysis(market_data_json)

    # Build chipset block (PCR + foreign futures)
    _chipset_parts: list[str] = []
    if not futures_df.empty:
        _fn = _safe_float(_ft_row.get("foreign_futures_net"))
        if _fn is not None:
            _chipset_parts.append(f"外資期貨 `{int(_fn):+,}口`")
    if not mkinst_df.empty:
        _fni = _safe_float(_mk_row.get("foreign_inst_norm"))
        if _fni is not None:
            _chipset_parts.append(f"現貨 `{_fni*100:+.0f}%`（標準化）")
    if _pcr_val is not None:
        _chipset_parts.append(f"PCR `{_pcr_val:.2f}`")
    chipset_block = ("💼 **籌碼**\n   " + " | ".join(_chipset_parts)) if _chipset_parts else ""

    scenario_block = ("🎯 **市場結構分析**\n   " + scenario_text.replace("\n", "\n   ")) if scenario_text else ""

    # ── TAIEX 技術指標 (RSI / MACD 直方 / 距MA60) ─────────────────────────────
    from indicators import add_rsi, add_macd
    _close = market_df["close"].astype(float)
    _taiex_tech: dict[str, float | None] = {"rsi14": None, "macd_hist": None, "dist_ma60": None}
    try:
        _rsi_s = add_rsi(_close, 14)
        _taiex_tech["rsi14"] = float(_rsi_s.iloc[-1]) if not _rsi_s.empty else None
    except Exception:
        pass
    try:
        _macd_df = add_macd(_close)
        _taiex_tech["macd_hist"] = float(_macd_df["macd_hist"].iloc[-1]) if not _macd_df.empty else None
    except Exception:
        pass
    try:
        if len(_close) >= 60:
            _ma60 = _close.rolling(60).mean().iloc[-1]
            _last_close = _close.iloc[-1]
            _taiex_tech["dist_ma60"] = (_last_close / _ma60 - 1) * 100 if _ma60 else None
    except Exception:
        pass

    _tech_parts: list[str] = []
    if _taiex_tech["rsi14"] is not None:
        _tech_parts.append(f"RSI `{_taiex_tech['rsi14']:.0f}`")
    if _taiex_tech["macd_hist"] is not None:
        _tech_parts.append(f"MACD直方 `{_taiex_tech['macd_hist']:+.1f}`")
    if _taiex_tech["dist_ma60"] is not None:
        _tech_parts.append(f"距MA60 `{_taiex_tech['dist_ma60']:+.1f}%`")
    taiex_block = ("📈 **加權技術**\n   " + " | ".join(_tech_parts)) if _tech_parts else ""

    # ── Claude 盤前解讀 ────────────────────────────────────────────────────────
    from claude_insight import generate_premarket_insight
    _pm_market = {
        "xgb_prob_up":  pred["prob_up"],
        "futures_net":  _safe_float(_ft_row.get("foreign_futures_net")),
        "vix":          _safe_float(_us_row.get("vix")),
        "night_change": night_data.get("change") if night_data else None,
        "night_trend":  night_data.get("last_hour_trend", "") if night_data else "",
        "cash_norm":    _safe_float(_mk_row.get("foreign_inst_norm")),
        "pcr":          _pcr_val,
        "nasdaq_ret":   _safe_float(_us_row.get("nasdaq_ret1")),
        "sox_ret":      _safe_float(_us_row.get("sox_ret1")),
        "tsm_adr_ret":  _safe_float(_us_row.get("tsm_adr_ret1")),
    }
    ai_insight = generate_premarket_insight(_pm_market, _taiex_tech)

    latest_date = pd.Timestamp(market_df["date"].max()).strftime("%Y-%m-%d")
    header = f"🌅 **盤前預測** · {latest_date} · {_cst_now()} CST"
    pred_block = format_prediction_block(
        pred,
        us_block=format_us_block(us_df if not us_df.empty else None),
        night_block=format_night_session_block(night_data),
        chipset_block=chipset_block,
        taiex_block=taiex_block,
        futures_block=format_futures_block(futures_df),
        scenario_block=scenario_block,
        news_block=format_market_news_block(market_news),
        calendar_block=format_calendar_block(calendar_events),
    )
    message = header + "\n\n" + (pred_block if pred_block else "⚠️ AI 模型訓練資料不足，無法產生預測。")
    if ai_insight:
        message += f"\n\n🤖 **今日操盤要點**\n{ai_insight}"

    # ── Market Regime Classification (新模組) ─────────────────────────────────────
    try:
        from market_regime.classifier import MarketRegimeClassifier
        from market_regime.scenario import generate_scenario
        _vix_val = float(_us_row.get("vix", 20) or 20)
        _fut_net = int(_safe_float(_ft_row.get("foreign_futures_net")) or 0)
        _night_chg = float(night_data.get("change", 0) if night_data else 0)
        _adv_ratio = 0.5  # breadth not available in predict mode; use neutral
        _tech_str = float(_safe_float(_us_row.get("nasdaq_ret1")) or 0) * 30 + \
                    float(_safe_float(_us_row.get("sox_ret1")) or 0) * 20
        _regime = MarketRegimeClassifier().classify(
            vix=_vix_val,
            futures_net=_fut_net,
            night_change=_night_chg,
            us_tech_strength=min(5, max(-5, _tech_str)),
            advance_ratio=_adv_ratio,
            xgb_prob_up=pred["prob_up"],
            pcr=_pcr_val or 1.0,
        )
        _scenario = generate_scenario(
            regime_label=_regime.label,
            regime_zh=_regime.label_zh,
            win_rate=_regime.win_rate_estimate,
            vix=_vix_val,
            futures_net=_fut_net,
            night_change=_night_chg,
            us_tech_strength=min(5, max(-5, _tech_str)),
            pcr=_pcr_val or 1.0,
            xgb_prob_up=pred["prob_up"],
        )
        message += f"\n\n{_scenario.format_discord()}"
    except Exception as _re_exc:
        _safe_print(f"[predict] 市場 regime 分類失敗（skip）: {_re_exc}")

    # ── Risk Assessment (新模組) ──────────────────────────────────────────────────
    try:
        from risk_engine.monitor import RiskMonitor
        _risk = RiskMonitor().assess(
            trade_date=today,
            vix=_vix_val,
            vix_change=float(_safe_float(_us_row.get("vix")) or 0) - float(_safe_float(_us_row.get("vix_prev")) or _vix_val),
            futures_net=_fut_net,
            night_change=_night_chg,
            advance_ratio=_adv_ratio,
            pcr=_pcr_val or 1.0,
            xgb_prob_up=pred["prob_up"],
            upcoming_events=calendar_events if calendar_events else [],
        )
        message += f"\n\n{_risk.format_discord()}"
    except Exception as _risk_exc:
        _safe_print(f"[predict] 風險評估失敗（skip）: {_risk_exc}")

    # ── News Sentiment (新模組) ───────────────────────────────────────────────────
    try:
        from news_engine.market_news import fetch_market_sentiment
        _news_result = fetch_market_sentiment(
            news_client=news_client if "news_client" in dir() else None,
            finmind_client=client,
            days=2,
        )
        _news_block = _news_result.format_discord()
        if _news_block:
            message += f"\n\n{_news_block}"
    except Exception as _ne_exc:
        _safe_print(f"[predict] 新聞情緒分析失敗（skip）: {_ne_exc}")

    # ── Save prediction JSON for web dashboard ────────────────────────────────
    try:
        import json as _json
        _pred_out = Path(args.output) / "prediction_latest.json"
        _pred_out.parent.mkdir(parents=True, exist_ok=True)
        _pred_payload: dict = {
            "date": today,
            "generated_at": _cst_now(),
            "xgb_prob_up": round(pred["prob_up"], 3),
            "xgb_label": pred.get("label", ""),
            "market_data": {
                "vix":         _safe_float(_us_row.get("vix")),
                "nasdaq_ret":  _safe_float(_us_row.get("nasdaq_ret1")),
                "sox_ret":     _safe_float(_us_row.get("sox_ret1")),
                "tsm_adr_ret": _safe_float(_us_row.get("tsm_adr_ret1")),
                "futures_net": _safe_float(_ft_row.get("foreign_futures_net")),
                "night_change": (night_data.get("change") if isinstance(night_data, dict) else None),
                "pcr":         (round(_pcr_val, 3) if _pcr_val is not None else None),
                "taiex_rsi":   _taiex_tech.get("rsi14"),
                "macd_hist":   _taiex_tech.get("macd_hist"),
                "dist_ma60":   _taiex_tech.get("dist_ma60"),
                "night_trend": (night_data.get("last_hour_trend") if isinstance(night_data, dict) else None),
            },
        }
        if ai_insight:
            _pred_payload["ai_insight"] = ai_insight
        if "_regime" in dir():
            _pred_payload["regime"] = {
                "label": getattr(_regime, "label", ""),
                "label_zh": getattr(_regime, "label_zh", ""),
                "win_rate": round(getattr(_regime, "win_rate_estimate", 0) * 100, 1),
            }
        if "_scenario" in dir():
            _pred_payload["scenario"] = {
                "main_scenario": getattr(_scenario, "main_scenario", ""),
                "best_strategy": getattr(_scenario, "best_strategy", ""),
                "danger_signals": list(getattr(_scenario, "danger_signals", [])),
                "forbidden_actions": list(getattr(_scenario, "forbidden_actions", [])),
            }
        if "_risk" in dir():
            _rl = getattr(_risk, "overall_level", None)
            _pred_payload["risk"] = {
                "level": (_rl.value if hasattr(_rl, "value") else str(_rl)) if _rl is not None else "MEDIUM",
                "score": float(getattr(_risk, "composite_score", 0)),
                "factors": [
                    {"name": f.name, "description": f.description, "action": f.action}
                    for f in getattr(_risk, "risk_factors", [])
                ],
            }
        if "_news_result" in dir():
            _pred_payload["news_sentiment"] = {
                "market_impact": float(getattr(_news_result, "market_impact_score", 0)),
                "key_events": list(getattr(_news_result, "key_events", [])),
                "bullish_count": int(getattr(_news_result, "bullish_count", 0)),
                "bearish_count": int(getattr(_news_result, "bearish_count", 0)),
            }
        _pred_out.write_text(_json.dumps(_pred_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _safe_print(f"[predict] 預測 JSON 已存至 {_pred_out}")
        # Append to prediction_history.json (keep last 90 days)
        _hist_out = _pred_out.parent / "prediction_history.json"
        try:
            _hist: list = _json.loads(_hist_out.read_text(encoding="utf-8")) if _hist_out.exists() else []
        except Exception:
            _hist = []
        _hist = [h for h in _hist if h.get("date", "") != today]  # remove same-day entry
        _hist.append(_pred_payload)
        _hist.sort(key=lambda h: h.get("date", ""), reverse=True)
        _hist = _hist[:90]
        _hist_out.write_text(_json.dumps(_hist, ensure_ascii=False, indent=2), encoding="utf-8")
        _safe_print(f"[predict] 歷史記錄已更新（{len(_hist)} 筆）→ {_hist_out}")
    except Exception as _pj_exc:
        _safe_print(f"[predict] 儲存預測 JSON 失敗（skip）: {_pj_exc}")

    _safe_print(message)
    if args.notify:
        send_discord_messages([message])


def _check_positions(
    prices: pd.DataFrame,
    positions_path: Path,
    notify: bool,
) -> None:
    """Check open positions against latest batch prices and alert via Discord."""
    if not positions_path.exists():
        return
    try:
        pos = pd.read_csv(positions_path, encoding="utf-8-sig")
    except Exception as exc:
        _safe_print(f"[positions] 讀取失敗：{exc}")
        return

    required = {"stock_id", "entry_price", "stop_loss", "target"}
    if not required.issubset(pos.columns):
        _safe_print(f"[positions] 欄位不足，需要：{required}")
        return

    pos["stock_id"] = pos["stock_id"].astype(str).str.strip()
    latest = prices[["stock_id", "close", "name"]].drop_duplicates("stock_id").set_index("stock_id")

    alerts: list[str] = []
    today = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8)))
    for _, row in pos.iterrows():
        sid = str(row["stock_id"])
        if sid not in latest.index:
            continue
        current = float(latest.at[sid, "close"])
        entry = float(row["entry_price"])
        stop = float(row["stop_loss"])
        target = float(row["target"])
        name = str(latest.at[sid, "name"]) if "name" in latest.columns else sid
        ret_pct = (current - entry) / entry * 100
        entry_date_str = str(row.get("entry_date", ""))
        days_held = 0
        if entry_date_str:
            try:
                days_held = (today.date() - datetime.strptime(entry_date_str, "%Y-%m-%d").date()).days
            except ValueError:
                pass

        if current <= stop:
            alerts.append(
                f"🔴 **停損警報** {sid} {name}\n"
                f"  現價 `{current:.2f}` ≤ 停損 `{stop:.2f}` | 報酬 `{ret_pct:+.1f}%` | 持倉 `{days_held}天`"
            )
        elif current >= target:
            alerts.append(
                f"🟢 **目標到達** {sid} {name}\n"
                f"  現價 `{current:.2f}` ≥ 目標 `{target:.2f}` | 報酬 `{ret_pct:+.1f}%` | 持倉 `{days_held}天`"
            )
        elif days_held > 20:
            alerts.append(
                f"🟡 **觀察提醒** {sid} {name}\n"
                f"  現價 `{current:.2f}` | 報酬 `{ret_pct:+.1f}%` | 持倉 `{days_held}天` | 目標 `{target:.2f}` 尚未到達"
            )

    if not alerts:
        _safe_print("[positions] 所有持倉正常，無警報")
        return

    header = f"📋 **持倉監控** · {today.strftime('%Y-%m-%d')}"
    msg = header + "\n\n" + "\n\n".join(alerts)
    _safe_print(msg)
    if notify:
        send_discord_messages(split_message(msg))


def run_partial_aggregate(args: argparse.Namespace) -> None:
    """Merge existing batch_NN.csv files and report current TOP N. Does NOT delete batch files."""
    batch_dir = Path(args.output) / "full_scan"
    csvs = sorted(batch_dir.glob("batch_*.csv")) if batch_dir.exists() else []
    if not csvs:
        msg = f"⚠️ **即時 TOP {args.top_n} 預覽失敗** · {args.end}\n找不到批次結果（{batch_dir}），請等掃描執行後再試。"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    frames = []
    for p in csvs:
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if not df.empty:
                frames.append(df)
        except Exception as exc:
            _safe_print(f"[partial-aggregate] 略過 {p.name}: {exc}")

    if not frames:
        msg = f"⚠️ **即時 TOP {args.top_n} 預覽** · {args.end}\n所有批次皆無候選。"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    all_candidates = pd.concat(frames, ignore_index=True)
    total_stocks = int(all_candidates["stock_id"].nunique()) if "stock_id" in all_candidates.columns else len(all_candidates)
    all_candidates = (
        all_candidates
        .sort_values("entry_score", ascending=False)
        .drop_duplicates(subset=["stock_id"])
        .reset_index(drop=True)
    )
    top = all_candidates.head(args.top_n)

    lines = [
        f"🔍 **即時 TOP {args.top_n} 預覽** · {args.end}",
        f"（掃描進行中）已掃 `{total_stocks}` 支 · {len(frames)} 批次",
        "",
    ]
    for rank, (_, row) in enumerate(top.iterrows(), 1):
        close = _float_or(row.get("close"))
        score = _float_or(row.get("entry_score"))
        cond  = _int_or(row.get("condition_count"))
        atr   = _float_or(row.get("atr14")) or None
        f_sc  = _int_or(row.get("f_score"), -1)
        f_tag = f" F`{f_sc}`" if f_sc >= 0 else ""
        _ind_raw = row.get("industry_category")
        industry = str(_ind_raw) if pd.notna(_ind_raw) and _ind_raw is not None else ""
        ind_tag = f" 〔{industry}〕" if industry else ""
        price_tag = _low_price_tag(row.get("close"))
        price_note = f" `{price_tag}`" if price_tag else ""
        entry_zone, stop, target, rr = _entry_stop_target(close, atr)
        _rsi = row.get("rsi14"); rsi_txt = f"RSI `{_float_or(_rsi):.1f}`" if pd.notna(_rsi) else "RSI `-`"
        _adx = row.get("adx14"); adx_txt = f"ADX `{_float_or(_adx):.1f}`" if pd.notna(_adx) else "ADX `-`"
        _vr  = row.get("volume_ratio"); vr_txt = f"量比 `{_float_or(_vr):.1f}x`" if pd.notna(_vr) else "量比 `-`"
        foreign_streak = _int_or(row.get("foreign_buy_streak"))
        invest_streak = _int_or(row.get("invest_trust_streak"))
        trend_inline = _trend_label(row)
        obs = recommend_observation_period(row, is_candidate=True)
        invest_txt = f" | 投信 `{invest_streak}d`" if invest_streak >= 1 else ""
        _margin_chg = _float_or(row.get("margin_change_5d"))
        _margin_tag = ""
        if _margin_chg < -3.0:
            _margin_tag = f" 📉融資 `-{abs(_margin_chg):.1f}%`"
        elif _margin_chg > 5.0:
            _margin_tag = f" ⚠️融資 `+{_margin_chg:.1f}%`"
        lines.append("━━━━━━━━━━━━━━━━━━━━")
        lines.append(
            f"**#{rank} {row.get('stock_id')}** {row.get('name', '')}{ind_tag}{price_note}{f_tag}"
            f"  分 `{score:.0f}` | `{cond}/{_MAX_CONDITION_COUNT}條`"
        )
        lines.append(f"💰 收 `{close:.2f}` | 進場 `{entry_zone}` | 停損 `{stop}` | 目標 `{target}` | R:R `{rr}`")
        lines.append(f"📊 {rsi_txt} | {adx_txt} | {vr_txt}{trend_inline}")
        lines.append(f"🏦 外資連買 `{foreign_streak}d`{invest_txt}{_margin_tag}")
        lines.append(f"⏱ {obs}")

    message = "\n".join(lines)
    _safe_print(message)

    # Save summary files (GitHub-renderable MD + downloadable XLSX)
    batch_dir.mkdir(parents=True, exist_ok=True)
    md_path = batch_dir / f"partial_top{args.top_n}.md"
    md_path.write_text(message, encoding="utf-8")
    _safe_print(f"[partial-aggregate] 摘要 → {md_path}")
    xlsx_path = batch_dir / f"partial_top{args.top_n}.xlsx"
    top.to_excel(xlsx_path, index=False, engine="openpyxl")
    _safe_print(f"[partial-aggregate] Excel → {xlsx_path}")

    if args.notify:
        send_discord_messages(split_message(message))


def _load_historical_top(
    batch_dir: "Path",
    today_str: str,
    past_days: int = 5,
    top_k: int = 50,
) -> "dict[str, list[dict]]":
    """Return {stock_id: [{date, score, rank}, ...]} for stocks in top_k over past_days."""
    import pandas as _pd
    history: dict[str, list[dict]] = {}
    today = _pd.Timestamp(today_str)
    for delta in range(1, past_days + 1):
        day = (today - _pd.Timedelta(days=delta)).strftime("%Y-%m-%d")
        day_files = sorted(batch_dir.glob(f"batch_seq*_{day}.csv"))
        if not day_files:
            continue
        frames = []
        for p in day_files:
            try:
                df = _pd.read_csv(p, encoding="utf-8-sig")
                if not df.empty and "entry_score" in df.columns and "stock_id" in df.columns:
                    frames.append(df)
            except Exception:
                continue
        if not frames:
            continue
        merged = (
            _pd.concat(frames, ignore_index=True)
            .sort_values("entry_score", ascending=False)
            .drop_duplicates(subset=["stock_id"])
        )
        for rank, (_, row) in enumerate(merged.head(top_k).iterrows(), 1):
            sid = str(row["stock_id"])
            if sid not in history:
                history[sid] = []
            history[sid].append({"date": day, "score": float(row["entry_score"]), "rank": rank})
    return history


def run_aggregate(args: argparse.Namespace) -> None:
    """Merge all batch_NN.csv files from output/full_scan/ and send top-N to Discord."""
    batch_dir = Path(args.output) / "full_scan"
    _all_csvs = sorted(batch_dir.glob("batch_*.csv")) if batch_dir.exists() else []
    _date_pat = re.compile(r"(\d{4}-\d{2}-\d{2})")
    _dates = [m.group(1) for f in _all_csvs if (m := _date_pat.search(f.name))]
    if _dates:
        _latest = max(_dates)
        csvs = [f for f in _all_csvs if _latest in f.name]
    else:
        csvs = _all_csvs
    if not csvs:
        msg = f"⚠️ **全市場彙整失敗** · {args.end}\n找不到批次結果（{batch_dir}），請確認掃描已執行。"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    frames = []
    empty_count = 0
    for p in csvs:
        try:
            df = pd.read_csv(p, encoding="utf-8-sig")
            if not df.empty:
                frames.append(df)
            else:
                empty_count += 1
                _safe_print(f"[aggregate] {p.name} 為空（本批無候選）")
        except Exception as exc:
            _safe_print(f"[aggregate] 略過 {p.name}: {exc}")

    if not frames:
        if empty_count > 0:
            msg = (
                f"⚠️ **全市場彙整：本日無候選** · {args.end}\n"
                f"共 {len(csvs)} 批次，{empty_count} 批無符合條件的股票（可能為 API 配額不足或市場條件不符）。"
            )
        else:
            msg = f"⚠️ **全市場彙整失敗** · {args.end}\n找不到可讀取的批次結果，請確認掃描已執行。"
        _safe_print(msg)
        if args.notify:
            send_discord_messages([msg])
        return

    all_candidates = pd.concat(frames, ignore_index=True)

    # ── 乾淨化 ──────────────────────────────────────────────────────────────
    # 1. 過濾超過 7 天的極舊資料（允許跨週末/假日的資料日期差異）
    if "date" in all_candidates.columns:
        all_candidates["date"] = pd.to_datetime(all_candidates["date"])
        _data_latest = all_candidates["date"].max()
        _cutoff = _data_latest - pd.Timedelta(days=7)
        _before = len(all_candidates)
        all_candidates = all_candidates[all_candidates["date"] >= _cutoff]
        if (_removed := _before - len(all_candidates)) > 0:
            _safe_print(f"[aggregate] 移除過期資料（7天前）{_removed} 筆")
    # 2. 移除收盤價為 0 / inf / NaN
    if "close" in all_candidates.columns:
        _close_num = pd.to_numeric(all_candidates["close"], errors="coerce")
        _before = len(all_candidates)
        all_candidates = all_candidates[
            _close_num.notna() & (_close_num > 0) & np.isfinite(_close_num)
        ]
        if (_removed := _before - len(all_candidates)) > 0:
            _safe_print(f"[aggregate] 移除價格異常（0/inf/NaN）資料 {_removed} 筆")

    total_stocks = int(all_candidates["stock_id"].nunique()) if "stock_id" in all_candidates.columns else len(all_candidates)
    all_candidates = (
        all_candidates
        .sort_values("entry_score", ascending=False)
        .drop_duplicates(subset=["stock_id"])
        .reset_index(drop=True)
    )
    top = all_candidates.head(args.top_n)

    lines = [
        f"🌙 **全市場掃描 TOP {args.top_n}** · {args.end}",
        f"共 {len(frames)} 批次 · 掃描 {total_stocks} 支",
        "",
    ]
    for rank, (_, row) in enumerate(top.iterrows(), 1):
        close = _float_or(row.get("close"))
        score = _float_or(row.get("entry_score"))
        cond  = _int_or(row.get("condition_count"))
        atr   = _float_or(row.get("atr14")) or None
        f_sc  = _int_or(row.get("f_score"), -1)
        f_tag = f" F`{f_sc}`" if f_sc >= 0 else ""
        _ind_raw = row.get("industry_category")
        industry = str(_ind_raw) if pd.notna(_ind_raw) and _ind_raw is not None else ""
        ind_tag = f" 〔{industry}〕" if industry else ""
        price_tag = _low_price_tag(row.get("close"))
        price_note = f" `{price_tag}`" if price_tag else ""
        entry_zone, stop, target, rr = _entry_stop_target(close, atr)
        _rsi = row.get("rsi14"); rsi_txt = f"RSI `{_float_or(_rsi):.1f}`" if pd.notna(_rsi) else "RSI `-`"
        _adx = row.get("adx14"); adx_txt = f"ADX `{_float_or(_adx):.1f}`" if pd.notna(_adx) else "ADX `-`"
        _vr  = row.get("volume_ratio"); vr_txt = f"量比 `{_float_or(_vr):.1f}x`" if pd.notna(_vr) else "量比 `-`"
        foreign_streak = _int_or(row.get("foreign_buy_streak"))
        invest_streak = _int_or(row.get("invest_trust_streak"))
        trend_inline = _trend_label(row)
        obs = recommend_observation_period(row, is_candidate=True)
        invest_txt = f" | 投信 `{invest_streak}d`" if invest_streak >= 1 else ""
        _margin_chg = _float_or(row.get("margin_change_5d"))
        _margin_tag = ""
        if _margin_chg < -3.0:
            _margin_tag = f" 📉融資 `-{abs(_margin_chg):.1f}%`"
        elif _margin_chg > 5.0:
            _margin_tag = f" ⚠️融資 `+{_margin_chg:.1f}%`"
        lines.append("━━━━━━━━━━━━━━━━━━━━")
        lines.append(
            f"**#{rank} {row.get('stock_id')}** {row.get('name', '')}{ind_tag}{price_note}{f_tag}"
            f"  分 `{score:.0f}` | `{cond}/{_MAX_CONDITION_COUNT}條`"
        )
        lines.append(f"💰 收 `{close:.2f}` | 進場 `{entry_zone}` | 停損 `{stop}` | 目標 `{target}` | R:R `{rr}`")
        lines.append(f"📊 {rsi_txt} | {adx_txt} | {vr_txt}{trend_inline}")
        lines.append(f"🏦 外資連買 `{foreign_streak}d`{invest_txt}{_margin_tag}")
        lines.append(f"⏱ {obs}")

    # ── 跨日持續強勢 ──────────────────────────────────────────────────────────────
    _hist = _load_historical_top(batch_dir, args.end, past_days=5, top_k=50)
    _today_top50 = all_candidates.head(50)
    _persistence_rows = []
    for _, row in _today_top50.iterrows():
        sid = str(row.get("stock_id", ""))
        hist_entries = _hist.get(sid, [])
        days_count = len(hist_entries)
        if days_count >= 2:
            prev_score = hist_entries[-1]["score"]
            today_score = float(row.get("entry_score", 0))
            _persistence_rows.append({
                "stock_id": sid,
                "name": row.get("name", ""),
                "today_score": today_score,
                "days_in_top": days_count + 1,
                "score_delta": today_score - prev_score,
            })
    _persistence_rows.sort(key=lambda x: (-x["days_in_top"], -x["today_score"]))
    if _persistence_rows:
        lines.append("")
        lines.append("📅 **連續入榜（近5天 TOP 50）**")
        for i, p in enumerate(_persistence_rows[:8], 1):
            arrow = "↑" if p["score_delta"] > 0 else ("↓" if p["score_delta"] < 0 else "→")
            delta_str = f" {arrow} `{abs(p['score_delta']):.0f}`" if p["score_delta"] != 0 else ""
            lines.append(
                f"   {i}. **{p['stock_id']}** {p['name']}  連續 `{p['days_in_top']}` 天 | 分 `{p['today_score']:.0f}`{delta_str}"
            )

    # ── 融資籌碼統計（TOP N 範圍）────────────────────────────────────────────
    if "margin_change_5d" in all_candidates.columns:
        _top_margin = top["margin_change_5d"].apply(_float_or) if "margin_change_5d" in top.columns else pd.Series(dtype=float)
        _clean_cnt = int((_top_margin < -3.0).sum())
        _surge_cnt = int((_top_margin > 5.0).sum())
        if _clean_cnt > 0 or _surge_cnt > 0:
            lines.append("")
            lines.append(f"📉 融資籌碼乾淨（5日縮>3%）：`{_clean_cnt}` 支 ｜ ⚠️ 融資暴增（>5%）：`{_surge_cnt}` 支")

    message = "\n".join(lines)

    # ── Claude AI 精選推薦 ────────────────────────────────────────────────────
    try:
        from claude_insight import generate_daily_picks as _gen_picks
        _ai_text = _gen_picks(top_stocks=top.to_dict("records"))
        if _ai_text:
            message += f"\n\n🤖 **AI 精選推薦**\n━━━━━━━━━━━━━━━━━━━━\n{_ai_text}"
    except Exception as _ai_exc:
        _safe_print(f"[Claude] AI 區塊略過: {_ai_exc}")

    _safe_print(message)
    if args.notify:
        send_discord_messages(split_message(message))

    # Check open positions against latest prices
    positions_csv = Path(getattr(args, "positions", None) or "positions.csv")
    _check_positions(all_candidates, positions_csv, getattr(args, "notify", False))

    # 產生彙整 Excel（所有掃描股票，含 TOP 20 / 候選進場 / 無訊號 標記）
    _top_ids_set = set(top["stock_id"].tolist()) if "stock_id" in top.columns else set()
    agg_df = all_candidates.copy()
    agg_df["類型"] = agg_df.apply(
        lambda r: "TOP 20" if str(r.get("stock_id", "")) in _top_ids_set
        else ("候選進場" if r.get("entry_signal", False) else "無訊號"),
        axis=1,
    )
    agg_xlsx = batch_dir / f"aggregate_{args.end}.xlsx"
    try:
        agg_df.to_excel(agg_xlsx, index=False, engine="openpyxl")
        _safe_print(f"[aggregate] 彙整 Excel 已產生：{agg_xlsx.name}（{len(agg_df)} 支股票）")
    except Exception as _xlsx_exc:
        _safe_print(f"[aggregate] 彙整 Excel 寫入失敗：{_xlsx_exc}")

    # Sync full scan results to Notion (all stocks; top N marked as "TOP 20")
    if notion_enabled():
        try:
            _top_ids = set(top["stock_id"].tolist()) if "stock_id" in top.columns else set()
            _ok, _fail = sync_scan_results(
                all_candidates, pd.DataFrame(), args.end,
                top_stock_ids=_top_ids,
            )
            _safe_print(f"[aggregate] Notion 同步完成，共 {len(all_candidates)} 筆（成功 {_ok}，失敗 {_fail}）")
            if args.notify and os.getenv("DISCORD_WEBHOOK_URL"):
                TS = _cst_now()
                if _ok > 0:
                    send_discord_messages([
                        f"✅ **Notion 同步完成** · {TS} CST\n"
                        f"成功 `{_ok}` 筆，失敗 `{_fail}` 筆（TOP 20 標記 `{len(_top_ids)}` 筆）"
                    ])
                else:
                    send_discord_messages([
                        f"⚠️ **Notion 同步失敗** · {TS} CST\n"
                        f"全部 `{_fail}` 筆寫入失敗（401 Unauthorized）\n"
                        f"請確認 GitHub Secrets 的 `NOTION_TOKEN` 是否為正確的 `secret_xxx...` 格式"
                    ])
        except Exception as _exc:
            _safe_print(f"[aggregate] Notion 同步失敗（不影響主流程）: {_exc}")
            if args.notify and os.getenv("DISCORD_WEBHOOK_URL"):
                TS = _cst_now()
                send_discord_messages([
                    f"⚠️ **Notion 同步失敗** · {TS} CST\n"
                    f"錯誤：`{_exc}`\n"
                    f"請確認：1) Notion database 已連結 Integration  2) NOTION_DATABASE_ID 正確"
                ])

    # ── 存 aggregate_latest.json（網頁讀取用，在刪 CSV 前）──────────────────────
    try:
        import json as _agg_json_lib

        def _to_py(v):
            if isinstance(v, (np.integer,)): return int(v)
            if isinstance(v, (np.floating,)): return None if (np.isnan(v) or np.isinf(v)) else float(v)
            if isinstance(v, float) and (v != v or v == float('inf') or v == float('-inf')): return None
            return v

        def _row_to_dict(row):
            return {k: _to_py(v) for k, v in row.items()}

        _ld_col = "limit_down_streak"
        _ld_alerts = []
        if _ld_col in all_candidates.columns:
            _ld_df = all_candidates[all_candidates[_ld_col].apply(lambda x: _float_or(x)) >= 3]
            _ld_alerts = [_row_to_dict(r) for _, r in _ld_df.sort_values(_ld_col, ascending=False).head(20).iterrows()]

        _entry_cnt = int((all_candidates["entry_signal"] == True).sum()) if "entry_signal" in all_candidates.columns else 0

        _agg_payload = {
            "date": args.end,
            "generated_at": _cst_now(),
            "total_scanned": total_stocks,
            "entry_count": _entry_cnt,
            "top_stocks": [_row_to_dict(r) for _, r in top.iterrows()],
            "limit_down_alerts": _ld_alerts,
            "persistent_strong": _persistence_rows[:20] if "_persistence_rows" in dir() else [],
            "margin_stats": {
                "clean_count": _clean_cnt if "_clean_cnt" in dir() else 0,
                "surge_count": _surge_cnt if "_surge_cnt" in dir() else 0,
            },
            "ai_picks_text": _ai_text if "_ai_text" in dir() else "",
        }
        _agg_out = Path(args.output) / "aggregate_latest.json"
        _agg_out.write_text(_agg_json_lib.dumps(_agg_payload, ensure_ascii=False, default=str), encoding="utf-8")
        _safe_print(f"[aggregate] 已存 aggregate_latest.json（{total_stocks} 支股票）")
    except Exception as _aj_exc:
        _safe_print(f"[aggregate] aggregate_latest.json 寫入失敗（無妨）: {_aj_exc}")

    # Clean up batch files after successful aggregation
    for p in csvs:
        try:
            p.unlink()
        except OSError:
            pass
    _safe_print(f"[aggregate] 清除 {len(csvs)} 個批次檔案")


def main() -> None:
    args = parse_args()
    cache_dir = Path(args.output) / "cache"
    client = FinMindClient(cache_dir=cache_dir)
    _cfg_overrides: dict[str, object] = {}
    if args.rsi_threshold is not None:
        _cfg_overrides["rsi_threshold"] = args.rsi_threshold
    if args.adx_threshold is not None:
        _cfg_overrides["adx_threshold"] = args.adx_threshold
    if args.stop_loss_pct is not None:
        _cfg_overrides["stop_loss_pct"] = args.stop_loss_pct
    if args.take_profit_pct is not None:
        _cfg_overrides["take_profit_pct"] = args.take_profit_pct
    if args.trailing_stop_pct is not None:
        _cfg_overrides["trailing_stop_pct"] = args.trailing_stop_pct
    if args.volume_multiplier is not None:
        _cfg_overrides["volume_multiplier"] = args.volume_multiplier
    if args.max_positions is not None:
        _cfg_overrides["max_positions"] = args.max_positions
    config = StrategyConfig(
        use_earnings_filter=args.use_earnings_filter,
        next_day_fill=args.next_day_fill,
        use_atr_stop=getattr(args, "use_atr_stop", False),
        atr_stop_multiplier=getattr(args, "atr_stop_multiplier", 2.0),
        max_holding_days=getattr(args, "max_holding_days", 0),
        max_positions_per_sector=getattr(args, "max_positions_per_sector", 2),
        f_score_min=getattr(args, "f_score_min", 0),
        **_cfg_overrides,
    )

    if getattr(args, "clean_cache", False):
        max_age = getattr(args, "clean_cache_days", 30)
        deleted = clean_cache(cache_dir, max_age_days=max_age)
        news_cache_dir = Path(args.output) / "news_cache"
        deleted += clean_cache(news_cache_dir, max_age_days=max_age)
        _safe_print(f"[cache] Deleted {deleted} stale cache files (data + news)")

    if args.mode == "backtest":
        run_backtest_mode(args, client, config)
    elif args.mode == "walk-forward":
        run_walk_forward(args, client, config)
    elif args.mode == "sponsor-monitor":
        run_sponsor_monitor(args, client, config)
    elif args.mode == "event-monitor":
        run_event_monitor(args, client, config)
    elif args.mode == "hybrid-monitor":
        run_hybrid_monitor(args, client, config)
    elif args.mode == "daily-report":
        run_daily_report(args, client, config)
    elif args.mode == "predict":
        run_predict(args, client, config)
    elif args.mode == "aggregate":
        run_aggregate(args)
    elif args.mode == "partial-aggregate":
        run_partial_aggregate(args)
    elif args.mode == "continue-scan":
        run_continue_scan(args, client, config)
    elif args.mode == "smart-scan":
        run_smart_scan(args, config)
    elif args.mode == "fill-batch-gaps":
        run_fill_batch_gaps(args, client, config)
    elif args.mode == "fill-gaps":
        run_fill_gaps(args, client, config)
    elif args.mode == "sequential-scan":
        run_sequential_scan(args, client, config)
    else:
        run_scan(args, client, config)


if __name__ == "__main__":
    main()
